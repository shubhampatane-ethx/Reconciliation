"""
=====================================================================
 DUMMY SERVER  ->  seed_utils.py
=====================================================================
NEW FILE — Multi-target addition.

Shared helpers for loading a target Excel workbook into a Postgres
table, exactly as TEXT columns. Extracted out of seed_cjbs_target.py
so that file and the new seed_targets.py (which seeds ALL registered
target files — see target_registry.py) share one identical, single
code path instead of two copies that could quietly drift apart.

Why TEXT for every column: the reconciliation engine (backend/normalize.py)
already normalizes every value — numbers, dates, text — for comparison on
both the Source and Target side, regardless of the underlying SQL type. So
there's no benefit to guessing stricter Postgres column types here, and
TEXT avoids type-conversion errors on values like "03-01-2026" (a date
that isn't in ISO format) or blank cells.
=====================================================================
"""

from datetime import datetime as _datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text


def stringify_cell(value):
    # Match how the main app's upload pipeline stringifies Excel cells
    # (pandas renders a midnight datetime as a bare date, e.g. "2026-03-01",
    # not "2026-03-01 00:00:00"). Keeping both sides in the same textual
    # format means a same-day match compares as identical rather than
    # showing up as a "Format Only" difference on every single row.
    if isinstance(value, _datetime):
        if value.time() == _datetime.min.time():
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    return str(value).strip()


def read_excel_rows(xlsx_path: Path):
    """
    Read a workbook's active sheet into (columns, rows), dropping any
    trailing "used range" columns whose header is blank and any fully
    blank trailing rows.
    """
    workbook = load_workbook(xlsx_path, data_only=True)
    worksheet = workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)
    raw_header = next(rows_iter)

    keep_idx = [
        i for i, h in enumerate(raw_header)
        if h is not None and str(h).strip() != ""
    ]
    columns = [str(raw_header[i]).strip() for i in keep_idx]

    rows = []
    for raw_row in rows_iter:
        values = [
            raw_row[i] if i < len(raw_row) else None
            for i in keep_idx
        ]
        if not any(v is not None and str(v).strip() != "" for v in values):
            continue  # skip fully blank trailing rows
        cleaned = [("" if v is None else stringify_cell(v)) for v in values]
        rows.append(cleaned)

    return columns, rows


def seed_table_from_excel(engine, xlsx_path: Path, table_name: str, label: str = None):
    """
    DROP + CREATE `table_name` (every column TEXT) from `xlsx_path`,
    then bulk-insert every row. Safe to re-run any time.

    Returns (row_count, column_count), or (0, 0) if the file doesn't
    exist (logged, not raised, so one missing workbook doesn't stop
    the other target tables from seeding).
    """
    tag = label or table_name
    if not xlsx_path.exists():
        print(f"[seed_utils] {tag}: Excel file not found at {xlsx_path} — skipping.")
        return 0, 0

    columns, rows = read_excel_rows(xlsx_path)
    print(f"[seed_utils] {tag}: loaded {len(rows)} row(s), {len(columns)} column(s) "
          f"from {xlsx_path.name}")

    quoted_col_defs = ", ".join(f'"{c}" TEXT' for c in columns)
    col_list = ", ".join(f'"{c}"' for c in columns)
    placeholders = ", ".join(f":c{i}" for i in range(len(columns)))

    with engine.begin() as conn:
        conn.execute(text(f'DROP TABLE IF EXISTS "{table_name}"'))
        conn.execute(text(f'CREATE TABLE "{table_name}" ({quoted_col_defs})'))

        insert_sql = text(f'INSERT INTO "{table_name}" ({col_list}) VALUES ({placeholders})')
        for row in rows:
            params = {f"c{i}": (row[i] if row[i] != "" else None) for i in range(len(columns))}
            conn.execute(insert_sql, params)

    print(f'[seed_utils] {tag}: done — "{table_name}" now has {len(rows)} row(s), '
          f'{len(columns)} column(s).')
    return len(rows), len(columns)
