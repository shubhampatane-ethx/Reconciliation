import json
import os
import uuid
from typing import Dict, List
from datetime import datetime
import re

BASE_DIR = os.path.dirname(__file__)
PERSIST_DIR = os.path.join(BASE_DIR, "vector_store")
META_FILE = os.path.join(PERSIST_DIR, "metadata.json")

os.makedirs(PERSIST_DIR, exist_ok=True)
REPORTS_DIR = os.path.join(PERSIST_DIR, "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Version-chain ("time series") storage
#
# Concept:
#   Source File (Day 0)  -> never changed, this is the baseline.
#   Day 1 File            -> compared against Day 0.
#   Day 2 File             -> compared against Day 1 (Day 1 becomes the new
#                              "source" for this comparison), and so on.
#
# Every uploaded version's actual data is kept on disk (as CSV) so that the
# *next* upload can be diffed against it. Only metadata + diff summaries are
# kept in series.json; the row-level diff report for each version transition
# is stored as its own JSON/XLSX file in REPORTS_DIR.
# ---------------------------------------------------------------------------
SERIES_FILE = os.path.join(PERSIST_DIR, "series.json")
SERIES_DATA_DIR = os.path.join(PERSIST_DIR, "series_data")
os.makedirs(SERIES_DATA_DIR, exist_ok=True)


def _load_metadata() -> Dict[str, Dict]:
    if not os.path.exists(META_FILE):
        return {}
    with open(META_FILE, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {}


def _save_metadata(meta: Dict[str, Dict]):
    with open(META_FILE, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2)


def _chunk_texts(texts: List[str], max_chars: int = 800) -> List[str]:
    chunks = []
    for text in texts:
        if len(text) <= max_chars:
            chunks.append(text)
            continue
        start = 0
        while start < len(text):
            chunks.append(text[start:start + max_chars])
            start += max_chars
    return chunks


def _extract_text_from_dataframe(df) -> List[str]:
    header = " | ".join(df.columns)
    rows = []
    for _, row in df.iterrows():
        row_text = "; ".join([f"{col}: {row[col]}" for col in df.columns])
        rows.append(f"{header}\n{row_text}")
    return _chunk_texts(rows)


def store_file(filename: str, df, file_type: str) -> Dict:
    file_id = uuid.uuid4().hex
    documents = _extract_text_from_dataframe(df)
    if not documents:
        documents = [f"{filename} ({file_type})"]

    metadata = _load_metadata()
    metadata[file_id] = {
        "file_id": file_id,
        "filename": filename,
        "file_type": file_type,
        "chunk_count": len(documents),
    }
    _save_metadata(metadata)

    chunks_file = os.path.join(PERSIST_DIR, f"{file_id}.json")
    with open(chunks_file, "w", encoding="utf-8") as f:
        json.dump(
            [{"chunk_index": idx, "text": text} for idx, text in enumerate(documents)],
            f,
            indent=2,
        )

    return metadata[file_id]


def list_files() -> List[Dict]:
    metadata = _load_metadata()
    return list(metadata.values())


def get_file_chunks(file_id: str) -> Dict:
    metadata = _load_metadata().get(file_id)
    if not metadata:
        return None

    chunks_file = os.path.join(PERSIST_DIR, f"{file_id}.json")
    if not os.path.exists(chunks_file):
        return {"metadata": metadata, "chunks": []}

    with open(chunks_file, "r", encoding="utf-8") as f:
        chunks = json.load(f)

    return {
        "metadata": metadata,
        "chunks": sorted(chunks, key=lambda c: c["chunk_index"]),
    }


def delete_file(file_id: str) -> bool:
    metadata = _load_metadata()
    if file_id not in metadata:
        return False

    metadata.pop(file_id)
    _save_metadata(metadata)

    chunks_file = os.path.join(PERSIST_DIR, f"{file_id}.json")
    if os.path.exists(chunks_file):
        os.remove(chunks_file)

    return True


def delete_all_files() -> int:
    """Remove every stored file (metadata + its chunk file). Returns the count."""
    metadata = _load_metadata()
    count = len(metadata)
    for file_id in list(metadata.keys()):
        chunks_file = os.path.join(PERSIST_DIR, f"{file_id}.json")
        if os.path.exists(chunks_file):
            os.remove(chunks_file)
    _save_metadata({})
    return count


def delete_all_reports() -> int:
    """Delete every generated Excel report. Returns the number removed."""
    count = 0
    for fn in os.listdir(REPORTS_DIR):
        if fn.endswith(".xlsx"):
            try:
                os.remove(os.path.join(REPORTS_DIR, fn))
                count += 1
            except OSError:
                pass
    return count


def _full_comparison_dataframe(rows: List[Dict], key_columns: List[str], before_label: str, after_label: str):
    """Every single row from BOTH files, side by side, in one table:
    matched rows, updated rows, deleted rows, and added rows all together,
    each tagged with a Status. This is the 'give me everything, not just a
    summary' sheet. Returns (dataframe, changed_columns_per_excel_row,
    status_per_excel_row) for highlighting."""
    import pandas as pd

    if not rows:
        return pd.DataFrame([{"Status": "None - no rows to show"}]), [], []

    # Determine the full column set from whichever row (source or target) has data.
    sample = next((r for r in rows if r.get("source_row") and r.get("target_row")), rows[0])
    sample_source = sample.get("source_row") or {}
    sample_target = sample.get("target_row") or {}
    # Fall back to any row that has a populated source/target if the first sample was one-sided.
    for r in rows:
        if r.get("source_row"):
            sample_source = {**r["source_row"], **sample_source}
        if r.get("target_row"):
            sample_target = {**r["target_row"], **sample_target}

    common_cols = [c for c in sample_source.keys() if c in sample_target and c not in key_columns]
    source_only_cols = [c for c in sample_source.keys() if c not in sample_target]
    target_only_cols = [c for c in sample_target.keys() if c not in sample_source]

    records = []
    changed_per_row = []
    status_per_row = []
    for row in rows:
        key = row.get("key", {})
        source_row = row.get("source_row") or {}
        target_row = row.get("target_row") or {}
        status = row.get("status", "")
        changed_columns = row.get("changed_columns") or []

        record = {"Status": status}
        record.update({f"Key: {k}": v for k, v in key.items()})
        for col in common_cols:
            record[f"{before_label} - {col}"] = source_row.get(col, "")
            record[f"{after_label} - {col}"] = target_row.get(col, "")
        for col in source_only_cols:
            record[f"{before_label} only - {col}"] = source_row.get(col, "")
        for col in target_only_cols:
            record[f"{after_label} only - {col}"] = target_row.get(col, "")

        records.append(record)
        changed_per_row.append(changed_columns)
        status_per_row.append(status)

    return pd.DataFrame(records), changed_per_row, status_per_row


def _highlight_full_comparison(writer, sheet_name: str, dataframe, changed_per_row: List[List[str]],
                                status_per_row: List[str], before_label: str, after_label: str):
    """Yellow-highlight the exact cells that changed on Updated/Format rows,
    and give Deleted/Added rows a light row-level tint so the status is
    visible at a glance while scanning the full table."""
    from openpyxl.styles import PatternFill

    if dataframe.empty:
        return
    worksheet = writer.sheets[sheet_name]
    cell_highlight = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
    deleted_tint = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
    added_tint = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
    columns = list(dataframe.columns)
    col_index = {name: idx + 1 for idx, name in enumerate(columns)}
    ncols = len(columns)

    for row_offset, (changed_columns, status) in enumerate(zip(changed_per_row, status_per_row)):
        excel_row = row_offset + 2
        if status == "Deleted":
            for col_idx in range(1, ncols + 1):
                worksheet.cell(row=excel_row, column=col_idx).fill = deleted_tint
        elif status == "Added":
            for col_idx in range(1, ncols + 1):
                worksheet.cell(row=excel_row, column=col_idx).fill = added_tint
        for col in changed_columns:
            before_col = f"{before_label} - {col}"
            after_col = f"{after_label} - {col}"
            if before_col in col_index:
                worksheet.cell(row=excel_row, column=col_index[before_col]).fill = cell_highlight
            if after_col in col_index:
                worksheet.cell(row=excel_row, column=col_index[after_col]).fill = cell_highlight


def _safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]\*:/\\?]", "_", name)
    return cleaned[:31] or "Sheet"


def _rows_dataframe(rows: List[Dict], status_label: str = None):
    """Full-row table (every original column) for missing/duplicate rows.
    A Status column is prepended so it's obvious *where* each row is
    missing from / why it's listed, without having to open the Summary tab."""
    import pandas as pd

    if not rows:
        return pd.DataFrame([{"Status": "None - no rows to show"}])

    df = pd.DataFrame(rows)
    # tidy up the internal reconciliation key column if present
    if "_reconciliation_key" in df.columns:
        df = df.drop(columns=["_reconciliation_key"])
    if status_label:
        df.insert(0, "Status", status_label)
    return df


def _issue_dataframe_wide(rows: List[Dict], key_columns: List[str], before_label: str, after_label: str):
    """One row per changed record. Shows:
      - Status column (Updated)
      - Key column(s)
      - Changed Columns (list of what moved)
      - For every CHANGED column only: Before | After side-by-side
      - Unchanged columns are shown once (as they are in both files)
    This keeps the sheet readable without hiding any data."""
    import pandas as pd

    if not rows:
        return pd.DataFrame([{"Status": "None - no rows to show"}]), []

    # Collect the superset of all changed column names across all rows.
    all_changed_cols_ordered = []
    seen = set()
    for row in rows:
        for col in (row.get("changed_columns") or [d["column"] for d in row.get("differences", [])]):
            if col not in seen:
                all_changed_cols_ordered.append(col)
                seen.add(col)

    # All non-key columns from source_row to show the unchanged context.
    sample_source = rows[0].get("source_row", {})
    unchanged_cols = [c for c in sample_source.keys()
                      if c not in key_columns and c not in seen]

    records = []
    changed_per_row = []
    for row in rows:
        key = row.get("key", {})
        source_row = row.get("source_row", {})
        target_row = row.get("target_row", {})
        changed_columns = row.get("changed_columns") or [d["column"] for d in row.get("differences", [])]

        record = {"Status": "Updated"}
        record.update({k: v for k, v in key.items()})
        record["Changed Columns"] = ", ".join(changed_columns) if changed_columns else "(format only)"
        # Changed columns: before | after (highlighted in caller)
        for col in all_changed_cols_ordered:
            record[f"{before_label} ← {col}"] = source_row.get(col, "")
            record[f"{after_label} → {col}"] = target_row.get(col, "")
        # Unchanged columns: just the source value (same in both)
        for col in unchanged_cols:
            record[col] = source_row.get(col, "")
        records.append(record)
        changed_per_row.append(changed_columns)

    return pd.DataFrame(records), changed_per_row


def _auto_size_columns(writer, sheet_name: str, dataframe):
    worksheet = writer.sheets[sheet_name]
    for idx, column in enumerate(dataframe.columns, start=1):
        values = dataframe[column].astype(str).tolist()
        max_length = max([len(str(column)), *(len(value) for value in values[:200])], default=12)
        worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = min(max_length + 2, 48)


def _style_header(writer, sheet_name: str, ncols: int, fill_hex: str = "1F4E78"):
    """Bold white header row with a colored fill, frozen header row, and an
    autofilter so the sheet behaves like a proper sortable/filterable table."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    worksheet = writer.sheets[sheet_name]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    for col_idx in range(1, ncols + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=False)
    worksheet.freeze_panes = "A2"
    last_col_letter = get_column_letter(ncols)
    last_row = worksheet.max_row
    if last_row >= 1 and ncols >= 1:
        worksheet.auto_filter.ref = f"A1:{last_col_letter}{max(last_row, 1)}"


def _highlight_changed_cells(writer, sheet_name: str, dataframe, changed_per_row: List[List[str]],
                              before_label: str, after_label: str):
    """Paint the exact Before/After cells that differ in yellow, so a reviewer
    can scan a wide table and immediately spot what moved."""
    from openpyxl.styles import PatternFill

    if dataframe.empty or not changed_per_row:
        return
    worksheet = writer.sheets[sheet_name]
    highlight = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
    columns = list(dataframe.columns)
    col_index = {name: idx + 1 for idx, name in enumerate(columns)}  # 1-based for openpyxl

    for row_offset, changed_columns in enumerate(changed_per_row):
        excel_row = row_offset + 2
        for col in changed_columns:
            # Try both naming conventions used in different sheets
            for before_col in [f"{before_label} ← {col}", f"{before_label} - {col}"]:
                if before_col in col_index:
                    worksheet.cell(row=excel_row, column=col_index[before_col]).fill = highlight
            for after_col in [f"{after_label} → {col}", f"{after_label} - {col}"]:
                if after_col in col_index:
                    worksheet.cell(row=excel_row, column=col_index[after_col]).fill = highlight


def _write_sheet(writer, sheet_name: str, dataframe, fill_hex: str = "1F4E78"):
    safe_sheet = _safe_sheet_name(sheet_name)
    dataframe.to_excel(writer, sheet_name=safe_sheet, index=False)
    _auto_size_columns(writer, safe_sheet, dataframe)
    _style_header(writer, safe_sheet, len(dataframe.columns), fill_hex=fill_hex)
    return safe_sheet


def store_report(report: Dict, source_meta: Dict, target_meta: Dict, key_columns: List[str], day_summary: List[Dict],
                  source_label: str = None, target_label: str = None) -> Dict:
    """Excel workbook — layout swapped to match the in-app renderFullComparison table exactly:
    The PRIMARY sheet 'All Rows' shows every row side-by-side:
        Status | Key | Source←col | Target→col for every column
    with yellow cell highlights on changed values, red row tints for Deleted,
    green row tints for Added — identical to what the user sees in the Discrepancies UI.

    Individual breakdown sheets (Deleted, Added, Updated, Format, Duplicates) now use
    the same flat Status-labelled layout that the UI shows when the user opens each
    details section.
    """
    import pandas as pd

    source_label = source_label or source_meta.get("filename", "Source")
    target_label = target_label or target_meta.get("filename", "Target")

    ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    safe_src = re.sub(r'[^a-zA-Z0-9_-]', '_', source_meta.get('filename', 'source'))
    safe_tgt = re.sub(r'[^a-zA-Z0-9_-]', '_', target_meta.get('filename', 'target'))
    fname = f"{ts}_{safe_src}_vs_{safe_tgt}_report.xlsx"
    path = os.path.join(REPORTS_DIR, fname)

    # ── Summary ───────────────────────────────────────────────────────────────
    total_full = report.get("full_comparison", {}).get("count", 0)
    matched = max(total_full
                  - report.get("mismatches", {}).get("count", 0)
                  - report.get("format_inconsistencies", {}).get("count", 0)
                  - report.get("missing_in_target", {}).get("count", 0)
                  - report.get("missing_in_source", {}).get("count", 0), 0)
    summary_df = pd.DataFrame([
        {"Metric": "Generated At (UTC)",          "Value": ts},
        {"Metric": "Source / Previous File",       "Value": source_label},
        {"Metric": "Target / New File",            "Value": target_label},
        {"Metric": "Key Columns",                  "Value": ", ".join(key_columns)},
        {"Metric": f"Rows in '{source_label}'",    "Value": report.get("source_record_count", 0)},
        {"Metric": f"Rows in '{target_label}'",    "Value": report.get("target_record_count", 0)},
        {"Metric": "Matched (No Change)",          "Value": matched},
        {"Metric": "Deleted",                      "Value": report.get("missing_in_target", {}).get("count", 0)},
        {"Metric": "Added",                        "Value": report.get("missing_in_source", {}).get("count", 0)},
        {"Metric": "Updated (Value Changes)",      "Value": report.get("mismatches", {}).get("count", 0)},
        {"Metric": "Renamed (Fuzzy Matched)",      "Value": report.get("fuzzy_matches", {}).get("count", 0)},
        {"Metric": "Format-Only Differences",      "Value": report.get("format_inconsistencies", {}).get("count", 0)},
        {"Metric": f"Duplicates in '{source_label}'", "Value": report.get("duplicates_source", {}).get("count", 0)},
        {"Metric": f"Duplicates in '{target_label}'", "Value": report.get("duplicates_target", {}).get("count", 0)},
    ])

    # ── Schema ────────────────────────────────────────────────────────────────
    schema = report.get("schema", {})
    schema_rows = (
        [{"Type": "Source-only column", "Column": c} for c in schema.get("source_only_columns", [])]
        + [{"Type": "Target-only column", "Column": c} for c in schema.get("target_only_columns", [])]
    )
    schema_df = (pd.DataFrame(schema_rows) if schema_rows
                 else pd.DataFrame([{"Status": "No schema differences — both files have the same columns"}]))

    # ── PRIMARY sheet helper: side-by-side (same as renderFullComparison) ─────
    # Every row from both files; Status | Key | Source←col | Target→col
    # This is what the user sees in the in-app "All rows, side by side" table.
    def _sidebyside_df(full_rows):
        if not full_rows:
            return pd.DataFrame([{"Status": "None — no rows to show"}]), [], []

        src_cols, tgt_cols = [], []
        for r in full_rows:
            for c in (r.get("source_row") or {}):
                if c not in src_cols: src_cols.append(c)
            for c in (r.get("target_row") or {}):
                if c not in tgt_cols: tgt_cols.append(c)
        common   = [c for c in src_cols if c in tgt_cols and c not in key_columns]
        src_only = [c for c in src_cols if c not in tgt_cols]
        tgt_only = [c for c in tgt_cols if c not in src_cols]

        records, changed_per, status_per = [], [], []
        for r in full_rows:
            src    = r.get("source_row") or {}
            tgt    = r.get("target_row") or {}
            key    = r.get("key", {})
            chg    = r.get("changed_columns") or []
            status = r.get("status", "")
            rec = {"Status": status}
            rec.update(key)
            for col in common:
                rec[f"{source_label} — {col}"] = src.get(col, "")
                rec[f"{target_label} — {col}"] = tgt.get(col, "")
            for col in src_only:
                rec[f"{source_label} only — {col}"] = src.get(col, "")
            for col in tgt_only:
                rec[f"{target_label} only — {col}"] = tgt.get(col, "")
            records.append(rec)
            changed_per.append(chg)
            status_per.append(status)
        return pd.DataFrame(records), changed_per, status_per

    # ── Breakdown sheet helper: flat full-row (same as renderRows in UI) ──────
    # Status | every original column — used for Deleted, Added, Duplicate sheets.
    def _flat_rows_df(raw_rows, status_label):
        if not raw_rows:
            return pd.DataFrame([{"Status": "None — no rows to show"}])
        clean = [{k: v for k, v in r.items() if k != "_reconciliation_key"} for r in raw_rows]
        df = pd.DataFrame(clean)
        df.insert(0, "Status", status_label)
        return df

    # ── Updated/Format sheet helper: matches renderIssueRows in UI ───────────
    # Status | Key | Changed Columns | Source←X | Target→X per CHANGED col | unchanged cols once
    def _issue_df(issue_rows, lbl_before, lbl_after):
        if not issue_rows:
            return pd.DataFrame([{"Status": "None — no rows to show"}]), []
        all_chg, seen = [], set()
        for r in issue_rows:
            for c in (r.get("changed_columns") or [d["column"] for d in r.get("differences", [])]):
                if c not in seen: all_chg.append(c); seen.add(c)
        sample_src = issue_rows[0].get("source_row", {})
        unchanged  = [c for c in sample_src if c not in key_columns and c not in seen]
        records, changed_per = [], []
        for r in issue_rows:
            src  = r.get("source_row", {})
            tgt  = r.get("target_row", {})
            chg  = r.get("changed_columns") or [d["column"] for d in r.get("differences", [])]
            rec  = {"Status": "Updated"}
            rec.update(r.get("key", {}))
            rec["Changed Columns"] = ", ".join(chg) or "(format only)"
            for col in all_chg:
                rec[f"{lbl_before} ← {col}"] = src.get(col, "")
                rec[f"{lbl_after} → {col}"]  = tgt.get(col, "")
            for col in unchanged:
                rec[col] = src.get(col, "")
            records.append(rec)
            changed_per.append(chg)
        return pd.DataFrame(records), changed_per

    # ── Renamed/fuzzy-matched sheet helper ─────────────────────────────────────
    # A row here means the key text didn't match exactly between files, but a
    # vector similarity search found a very likely same-record match anyway
    # (e.g. "Alpha Proj" -> "Project Alpha") instead of it showing up as a
    # false Deleted + Added pair. Uses the same "{label} ← col" / "{label} → col"
    # naming as _issue_df so the existing _highlight_changed_cells works as-is.
    def _fuzzy_df(fuzzy_rows, lbl_before, lbl_after):
        if not fuzzy_rows:
            return pd.DataFrame([{"Status": "None — no fuzzy-matched renames found"}]), []
        all_chg, seen = [], set()
        for r in fuzzy_rows:
            for c in (r.get("changed_columns") or []):
                if c not in seen: all_chg.append(c); seen.add(c)
        sample_src = fuzzy_rows[0].get("source_row", {})
        unchanged = [c for c in sample_src if c not in key_columns and c not in seen]
        records, changed_per = [], []
        for r in fuzzy_rows:
            src = r.get("source_row", {})
            tgt = r.get("target_row", {})
            chg = r.get("changed_columns") or []
            rec = {"Status": "Renamed", "Match Confidence": r.get("confidence", "")}
            for col in key_columns:
                rec[f"Key Before ({col})"] = r.get("key_before", {}).get(col, "")
                rec[f"Key After ({col})"] = r.get("key_after", {}).get(col, "")
            rec["Changed Columns"] = ", ".join(chg) or "(key only)"
            for col in all_chg:
                rec[f"{lbl_before} ← {col}"] = src.get(col, "")
                rec[f"{lbl_after} → {col}"] = tgt.get(col, "")
            for col in unchanged:
                rec[col] = src.get(col, "")
            records.append(rec)
            changed_per.append(chg)
        return pd.DataFrame(records), changed_per

    # ── Build all dataframes ──────────────────────────────────────────────────
    all_df, all_changed, all_status = _sidebyside_df(
        report.get("full_comparison", {}).get("rows", []))

    deleted_df  = _flat_rows_df(report.get("missing_in_target", {}).get("rows", []),
                                 f"Deleted — in '{source_label}', missing from '{target_label}'")
    added_df    = _flat_rows_df(report.get("missing_in_source", {}).get("rows", []),
                                 f"Added — new in '{target_label}', not in '{source_label}'")
    dup_src_df  = _flat_rows_df(report.get("duplicates_source", {}).get("rows", []),
                                 f"Duplicate key in '{source_label}'")
    dup_tgt_df  = _flat_rows_df(report.get("duplicates_target", {}).get("rows", []),
                                 f"Duplicate key in '{target_label}'")

    updated_df, updated_changed = _issue_df(
        report.get("mismatches", {}).get("rows", []), source_label, target_label)
    format_df,  format_changed  = _issue_df(
        report.get("format_inconsistencies", {}).get("rows", []), source_label, target_label)
    fuzzy_df, fuzzy_changed = _fuzzy_df(
        report.get("fuzzy_matches", {}).get("rows", []), source_label, target_label)

    # ── Highlight helper for side-by-side sheet ───────────────────────────────
    def _hl_sidebyside(writer, sheet_name, dataframe, changed_per, status_per):
        from openpyxl.styles import PatternFill
        if dataframe.empty: return
        ws   = writer.sheets[sheet_name]
        hl   = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
        del_ = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        add_ = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
        ren_ = PatternFill(start_color="F4ECF7", end_color="F4ECF7", fill_type="solid")
        cols = {name: i + 1 for i, name in enumerate(dataframe.columns)}
        ncols = len(dataframe.columns)
        for offset, (chg, status) in enumerate(zip(changed_per, status_per)):
            row = offset + 2
            if status == "Deleted":
                for c in range(1, ncols + 1): ws.cell(row=row, column=c).fill = del_
            elif status == "Added":
                for c in range(1, ncols + 1): ws.cell(row=row, column=c).fill = add_
            elif status == "Renamed":
                for c in range(1, ncols + 1): ws.cell(row=row, column=c).fill = ren_
            for col in chg:
                for label in [f"{source_label} — {col}", f"{target_label} — {col}"]:
                    if label in cols:
                        ws.cell(row=row, column=cols[label]).fill = hl

    # ── Write workbook ────────────────────────────────────────────────────────
    # Every sheet's ROW DATA is written FIRST. The (optional) cell highlighting
    # is applied afterwards as best-effort — a highlighting failure on one sheet
    # must never abort the workbook and leave a summary-only report behind. This
    # guarantees the Deleted / Added / Updated / Format / Duplicate rows are
    # always present in every downloaded report.
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_sheet(writer, "Summary",            summary_df,  fill_hex="1F4E78")
        if day_summary:
            day_df = pd.DataFrame(day_summary)
            _write_sheet(writer, "Day Wise", day_df, fill_hex="1F4E78")

        # PRIMARY sheet — side-by-side, mirrors the in-app "All rows" table
        all_sheet = _write_sheet(writer, "All Rows (Side by Side)", all_df, fill_hex="34495E")
        # Breakdown sheets — flat full-row, mirrors the in-app per-section tables
        _write_sheet(writer, "Deleted Rows",       deleted_df,  fill_hex="C0392B")
        _write_sheet(writer, "Added Rows",         added_df,    fill_hex="1E8449")
        upd_sheet = _write_sheet(writer, "Updated Rows", updated_df, fill_hex="B7950B")
        fmt_sheet = _write_sheet(writer, "Format Differences", format_df, fill_hex="7D6608")
        fzy_sheet = _write_sheet(writer, "Renamed (Fuzzy Matched)", fuzzy_df, fill_hex="884EA0")
        _write_sheet(writer, "Source Duplicates",  dup_src_df,  fill_hex="6C3483")
        _write_sheet(writer, "Target Duplicates",  dup_tgt_df,  fill_hex="6C3483")
        _write_sheet(writer, "Schema Differences", schema_df,   fill_hex="273746")

        # Best-effort highlighting (never allowed to drop a data sheet).
        for _hl in (
            lambda: _hl_sidebyside(writer, all_sheet, all_df, all_changed, all_status),
            lambda: _highlight_changed_cells(writer, upd_sheet, updated_df, updated_changed, source_label, target_label),
            lambda: _highlight_changed_cells(writer, fmt_sheet, format_df, format_changed, source_label, target_label),
            lambda: _highlight_changed_cells(writer, fzy_sheet, fuzzy_df, fuzzy_changed, source_label, target_label),
        ):
            try:
                _hl()
            except Exception:
                pass

    return {"report_file": fname, "path": path, "timestamp": ts}


def list_reports() -> List[Dict]:
    files = []
    for fn in sorted(os.listdir(REPORTS_DIR), reverse=True):
        if fn.endswith('.xlsx'):
            parts = fn.split('_')
            ts = parts[0] if parts else ''
            files.append({"filename": fn, "timestamp": ts, "format": "excel"})
    return files


# ---------------------------------------------------------------------------
# Version-chain ("time series") helpers
# ---------------------------------------------------------------------------

def _load_series_all() -> Dict[str, Dict]:
    if not os.path.exists(SERIES_FILE):
        return {}
    with open(SERIES_FILE, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {}


def _save_series_all(data: Dict[str, Dict]):
    with open(SERIES_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def _version_data_path(series_id: str, version: int) -> str:
    return os.path.join(SERIES_DATA_DIR, f"{series_id}_v{version}.csv")


def _save_version_dataframe(series_id: str, version: int, df):
    df.to_csv(_version_data_path(series_id, version), index=False)


def load_version_dataframe(series_id: str, version: int):
    """Load a previously stored version's data so it can be diffed again."""
    import pandas as pd

    path = _version_data_path(series_id, version)
    if not os.path.exists(path):
        return None
    return pd.read_csv(path, dtype=str).fillna("")


def create_series(name: str, filename: str, df) -> Dict:
    """Register a new series with the untouched Source file as Version 0."""
    all_series = _load_series_all()
    series_id = uuid.uuid4().hex
    now = datetime.utcnow().isoformat() + "Z"

    version_entry = {
        "version": 0,
        "label": "Source",
        "filename": filename,
        "uploaded_at": now,
        "row_count": int(len(df)),
        "column_count": int(len(df.columns)),
        "key_columns": None,
        "diff_summary": None,   # Version 0 is the baseline, nothing to diff against.
        "report_file": None,
    }
    all_series[series_id] = {
        "series_id": series_id,
        "name": name or filename,
        "created_at": now,
        "versions": [version_entry],
    }
    _save_series_all(all_series)
    _save_version_dataframe(series_id, 0, df)
    return all_series[series_id]


def list_series() -> List[Dict]:
    all_series = _load_series_all()
    result = []
    for s in all_series.values():
        versions = s.get("versions", [])
        result.append({
            "series_id": s["series_id"],
            "name": s["name"],
            "created_at": s["created_at"],
            "version_count": len(versions),
            "latest_version": versions[-1] if versions else None,
        })
    return result


def get_series(series_id: str) -> Dict:
    return _load_series_all().get(series_id)


def get_latest_version_number(series_id: str) -> int:
    s = get_series(series_id)
    versions = (s or {}).get("versions", [])
    return versions[-1]["version"] if versions else -1


def add_series_version(series_id: str, filename: str, df, key_columns: List[str],
                        diff_summary: Dict, report_file: str, label: str = None) -> Dict:
    """Append a new Day N version to the chain, right after diffing it
    against the previous version (Day N-1)."""
    all_series = _load_series_all()
    s = all_series.get(series_id)
    if not s:
        return None

    next_version = s["versions"][-1]["version"] + 1
    now = datetime.utcnow().isoformat() + "Z"
    version_entry = {
        "version": next_version,
        "label": label or f"Day {next_version}",
        "filename": filename,
        "uploaded_at": now,
        "row_count": int(len(df)),
        "column_count": int(len(df.columns)),
        "key_columns": key_columns,
        "diff_summary": diff_summary,
        "report_file": report_file,
    }
    s["versions"].append(version_entry)
    _save_series_all(all_series)
    _save_version_dataframe(series_id, next_version, df)
    return version_entry


def delete_series(series_id: str) -> bool:
    all_series = _load_series_all()
    s = all_series.pop(series_id, None)
    if not s:
        return False
    _save_series_all(all_series)
    for version_entry in s.get("versions", []):
        path = _version_data_path(series_id, version_entry["version"])
        if os.path.exists(path):
            os.remove(path)
    return True


def save_series_diff_json(series_id: str, version: int, diff_report: Dict) -> str:
    """Persist the full row-level diff between version-1 and version as JSON."""
    fname = f"series_{series_id}_v{version}_diff.json"
    path = os.path.join(REPORTS_DIR, fname)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(diff_report, f, indent=2)
    return fname


def store_series_excel_report(series_id: str, series_name: str, prev_label: str, curr_label: str,
                               version: int, diff_report: Dict, key_columns: List[str],
                               day_summary: List[Dict]) -> Dict:
    """Build the same rich Excel workbook used by one-off reconciliation,
    but named for this series' version transition (prev day -> this day)."""
    source_meta = {"filename": f"{series_name} - {prev_label}"}
    target_meta = {"filename": f"{series_name} - {curr_label}"}
    report_info = store_report(diff_report, source_meta, target_meta, key_columns, day_summary,
                                source_label=prev_label, target_label=curr_label)

    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', series_name)
    desired_name = f"series_{safe_name}_v{version}_{report_info['timestamp']}.xlsx"
    old_path = report_info["path"]
    new_path = os.path.join(REPORTS_DIR, desired_name)
    try:
        os.replace(old_path, new_path)
        report_info["report_file"] = desired_name
        report_info["path"] = new_path
    except OSError:
        pass  # keep the original name if rename fails for any reason
    return report_info


def load_series_diff_json(series_id: str, version: int) -> Dict:
    fname = f"series_{series_id}_v{version}_diff.json"
    path = os.path.join(REPORTS_DIR, fname)
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)