import os
import json
import functools
import difflib
from collections import Counter
from decimal import Decimal, InvalidOperation
from flask import Flask, g, request, jsonify
from flask_cors import CORS
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

from storage import (
    delete_file, store_file, list_files, get_file_chunks, store_report, list_reports, get_report_owner,
    create_series, list_series, list_series_for_user, get_series, add_series_version,
    delete_series, delete_all_series, load_version_dataframe, save_series_diff_json,
    load_series_diff_json, store_series_excel_report, delete_all_files, delete_all_reports,
)
from fuzzy_match import find_fuzzy_matches
from insights import generate_plain_english_summary
from ollama_service import generate_response, OllamaError
from groq_service import generate_response as groq_generate, GroqError
import db
from schema_engine import generate_schema_mapping_analysis
from row_reconcile_engine import get_row_previews, reconcile_by_row_indexing, auto_match_rows_by_keys
from auth import configure_jwt, require_auth, optional_auth, admin_required, auth_bp, ensure_admin_bootstrap
from admin_routes import admin_bp
from flask import send_file

try:
    from kafka_service import publish_recon_job, publish_audit_event, publish_erp_sync_event, publish_notification_event
except Exception as _kafka_err:
    print(f"[Kafka] Warning: kafka_service could not be loaded: {_kafka_err}")
    publish_recon_job = lambda *a, **k: False
    publish_audit_event = lambda *a, **k: False
    publish_erp_sync_event = lambda *a, **k: False
    publish_notification_event = lambda *a, **k: False


app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)

# Initialise JWT (reads JWT_SECRET from env, configures Flask-JWT-Extended)
configure_jwt(app)

# Register auth routes (/api/auth/register, /api/auth/login,
# /api/auth/refresh, /api/auth/logout, /api/auth/me)
app.register_blueprint(auth_bp)

# Register admin-only routes (/api/admin/...) — system-wide views across
# every user's data. Every route in admin_bp is protected by
# @admin_required, so mounting the blueprint is safe regardless of who
# else is logged in.
app.register_blueprint(admin_bp)

# All Postgres schema (users, series, datasets, series_versions,
# series_row_values, sessions) is created/migrated exclusively via
# Alembic — entrypoint.sh runs `alembic upgrade head` before this app
# starts. db.init_schema() no longer creates anything; it just verifies
# the tables it needs are present and logs a clear warning if
# migrations haven't been run yet, rather than creating/altering
# tables itself.
db.init_schema()

# Idempotently create the permanent Global Admin account
# (admin@gmail.com) if it doesn't already exist yet. No-op on every
# subsequent restart. See auth.ensure_admin_bootstrap for details.
ensure_admin_bootstrap()

ALLOWED_EXTENSIONS = {"csv", "xlsx", "xls"}
DATE_COLUMNS = [
    "date", "Date", "DATE", "transaction_date", "TransactionDate",
    "created_at", "CreatedAt",
]


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def _read_excel_unmerged(content: bytes) -> pd.DataFrame:
    """Read the first sheet of an .xlsx/.xls upload, unmerging ALL merged cells
    (including header-row merges) so every cell carries the value a person
    sees in Excel rather than blanks."""
    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active

    # FIX: removed the `if min_row <= 1: continue` guard so merged header
    # cells are also unmerged — previously they were skipped, leaving header
    # cells as None which became "Unnamed: N" column names.
    for merged_range in list(worksheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = worksheet.cell(row=min_row, column=min_col).value
        worksheet.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                worksheet.cell(row=row, column=col).value = top_left_value

    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return pd.DataFrame()

    columns = []
    seen = {}
    for idx, value in enumerate(header_row):
        name = str(value).strip() if value not in (None, "") else f"Unnamed: {idx}"
        if name in seen:
            seen[name] += 1
            name = f"{name}.{seen[name]}"
        else:
            seen[name] = 0
        columns.append(name)

    return pd.DataFrame(rows_iter, columns=columns)


def _drop_empty_unnamed_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Drop columns that have no header name AND are effectively empty.
    FIX: also drops Unnamed columns where >90% of values are blank — catches
    stray Excel formatting columns that have one or two accidental cell values."""
    cols_to_drop = []
    for col in df.columns:
        if str(col).startswith("Unnamed:"):
            series = df[col].astype(str).str.strip()
            non_empty = series[series != ""].shape[0]
            # Drop if entirely empty OR if less than 10% of rows have real data
            if non_empty == 0 or (len(series) > 0 and non_empty / len(series) < 0.1):
                cols_to_drop.append(col)
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
    return df


def read_dataframe(file_storage):
    filename = file_storage.filename
    ext = filename.rsplit('.', 1)[1].lower()
    content = file_storage.read()
    if ext == 'csv':
        df = pd.read_csv(BytesIO(content), dtype=str)
    elif ext == 'xls':
        df = pd.read_excel(BytesIO(content), dtype=str)
    else:
        df = _read_excel_unmerged(content)
    return _drop_empty_unnamed_columns(df)


def normalize_dataframe(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    df = df.fillna("")
    for col in df.columns:
        # FIX: convert via a lambda instead of bulk .astype(str) so that
        # native numeric types from openpyxl (int/float) are formatted cleanly:
        # integers stay as "1000" not "1000.0", and floats are rounded to 10
        # significant digits to avoid floating-point noise like "3.1400000000000001".
        def _to_str(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            if isinstance(v, float):
                # Remove trailing zeros: 3.14000...01 -> "3.14"
                return f"{v:.10g}"
            if isinstance(v, int):
                return str(v)
            return str(v).strip()
        df[col] = df[col].map(_to_str)
    return df


import re as _colname_re


def normalize_col_name(name):
    """Collapse a column name down to just its lowercase letters/digits, e.g.
    'PartyNumber' -> 'partynumber' and 'PARTY_NUMBER' -> 'partynumber'.
    Used only to DETECT that two differently-styled column names refer to
    the same field -- never used as the column name itself.
    """
    return _colname_re.sub(r'[^a-z0-9]', '', str(name).lower())


def align_equivalent_columns(df_source, df_target):
    """
    GENERIC fix for every op-co, not a one-off file edit: real-world Source
    and Target files for the SAME field are often named differently between
    systems -- e.g. an enrichment/validation export calls it 'PartyNumber'
    while the raw ERP export calls it 'PARTY_NUMBER'. Exact-name matching
    (what guess_key_columns/difference_summary use) treats those as two
    unrelated columns, so the key never gets detected and nothing lines up,
    no matter which op-co (Etairos, Airetech, ATS, ...) the files are for.

    This renames Target's column to match Source's column name whenever
    their normalized (lowercase, letters/digits-only) forms are identical
    AND the match is unambiguous (exactly one column on each side maps to
    that normalized form) AND they aren't already named identically (no
    pointless no-op rename). It only ever touches Target's naming to align
    it onto Source's -- values are never modified, only which column header
    a field is filed under.

    Returns (df_target_aligned, list_of_renames) where each rename is
    {"source_name": ..., "target_name_before": ...} for surfacing in the
    schema-differences report if useful.
    """
    source_norm = {}
    for col in df_source.columns:
        source_norm.setdefault(normalize_col_name(col), []).append(col)
    target_norm = {}
    for col in df_target.columns:
        target_norm.setdefault(normalize_col_name(col), []).append(col)

    rename_map = {}
    renames = []
    for norm, target_cols in target_norm.items():
        if len(target_cols) != 1:
            continue  # ambiguous on Target's side -- don't guess
        source_cols = source_norm.get(norm)
        if not source_cols or len(source_cols) != 1:
            continue  # no match, or ambiguous on Source's side
        target_col = target_cols[0]
        source_col = source_cols[0]
        if target_col == source_col:
            continue  # already identical -- nothing to align
        rename_map[target_col] = source_col
        renames.append({"source_name": source_col, "target_name_before": target_col})

    if rename_map:
        df_target = df_target.rename(columns=rename_map)
    return df_target, renames


def apply_manual_schema_mapping(df_target, schema_mapping):
    """
    Rename Target's columns onto Source's naming using an EXPLICIT mapping
    the user configured in the Schema Mapping modal:
    { source_column_name: target_column_name, ... }.

    This is what actually makes "Save & Reconcile" in that modal do
    anything -- without it, whatever the user picked in the modal was
    thrown away and the app silently fell back to the automatic
    align_equivalent_columns() name-normalisation heuristic, which finds
    nothing when Source/Target column names don't share a normalised form
    (e.g. an ERP export column with no resemblance to the enrichment
    file's column). Manual mapping always wins over the heuristic for the
    columns it covers; align_equivalent_columns() still runs afterward to
    pick up anything the user left unmapped.

    Entries mapped to '__ignore__' / blank are skipped (that source column
    intentionally has no target counterpart). Returns
    (df_target_renamed, list_of_renames).
    """
    if not schema_mapping:
        return df_target, []
    rename_map = {}
    renames = []
    used_targets = set()
    for source_col, target_col in schema_mapping.items():
        if not target_col or target_col in ('__ignore__', '-- Ignore / Skip --'):
            continue
        if target_col not in df_target.columns or target_col in used_targets:
            continue
        if target_col == source_col:
            continue
        rename_map[target_col] = source_col
        used_targets.add(target_col)
        renames.append({"source_name": source_col, "target_name_before": target_col})
    if rename_map:
        df_target = df_target.rename(columns=rename_map)
    return df_target, renames


def _parse_schema_mapping_form(form):
    """Parse the 'schema_mapping' JSON form field sent by the Schema Mapping
    modal. Returns {} on missing/invalid input rather than erroring the
    whole request -- a bad mapping should degrade to auto-alignment, not
    block the upload."""
    raw = (form.get('schema_mapping') or '').strip()
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except (ValueError, TypeError):
        return {}


def explicitly_ignored_source_columns(schema_mapping):
    """Source columns the user explicitly set to '-- Ignore / Skip --' in
    the Schema Mapping modal -- these must stay unmapped even during the
    fuzzy fallback pass below, since that was a deliberate choice."""
    return {
        sc for sc, tc in (schema_mapping or {}).items()
        if tc in ('__ignore__', '-- Ignore / Skip --')
    }


def _fuzzy_col_score(source_norm: str, target_norm: str) -> float:
    ratio = difflib.SequenceMatcher(None, source_norm, target_norm).ratio()
    # Boost when one is a clean substring of the other -- e.g. Source's
    # 'address1original' (from 'Address1_original') plainly contains
    # Target's 'address1' ('ADDRESS1'), which SequenceMatcher's raw ratio
    # alone doesn't score highly enough to clear the cutoff.
    if len(target_norm) >= 3 and (target_norm in source_norm or source_norm in target_norm):
        ratio = max(ratio, 0.85)
    return ratio


def fuzzy_align_remaining_columns(df_source, df_target, excluded_source_cols=None, cutoff: float = 0.72):
    """
    Third-tier column alignment, run AFTER manual mapping (exact, user
    -picked) and align_equivalent_columns() (exact normalized match) have
    both had their turn. Real-world Source extracts often carry
    suffixed/prefixed field names -- e.g. 'Address1_original',
    'PhoneNumber1_original' from an enrichment/staging export -- that don't
    *exactly* normalise to a match against a plainly-named Target column
    ('ADDRESS1'), so without this step those columns are left permanently
    "-- Ignored --" even though there's an obvious best-guess Target match.
    Only the primary/compare key needs to be picked deliberately; everything
    else the user hasn't explicitly mapped or explicitly skipped gets this
    best-effort auto-match, same as the Schema Mapping modal's own
    "AI Auto-Map" heuristic.

    For every Source column still without a same-named Target counterpart
    (and not in `excluded_source_cols`, i.e. not explicitly set to
    "-- Ignore / Skip --" by the user), scores it against every still
    -unclaimed Target column using a similarity ratio with a substring
    -containment boost, then does a greedy best-score-first, strictly
    one-to-one assignment for pairs clearing `cutoff`.

    Returns (df_target_aligned, list_of_fuzzy_renames).
    """
    excluded_source_cols = excluded_source_cols or set()
    source_cols = list(df_source.columns)
    target_cols = list(df_target.columns)

    unmatched_source = [c for c in source_cols if c not in target_cols and c not in excluded_source_cols]
    unmatched_target = [c for c in target_cols if c not in source_cols]
    if not unmatched_source or not unmatched_target:
        return df_target, []

    scored = []
    for sc in unmatched_source:
        sc_norm = normalize_col_name(sc)
        if not sc_norm:
            continue
        for tc in unmatched_target:
            tc_norm = normalize_col_name(tc)
            if not tc_norm:
                continue
            score = _fuzzy_col_score(sc_norm, tc_norm)
            if score >= cutoff:
                scored.append((score, sc, tc))

    # Greedy best-score-first assignment, strictly one-to-one on both sides
    # -- avoids two Source columns both grabbing the same Target column.
    scored.sort(key=lambda t: t[0], reverse=True)
    used_source, used_target = set(), set()
    rename_map = {}
    renames = []
    for score, sc, tc in scored:
        if sc in used_source or tc in used_target:
            continue
        rename_map[tc] = sc
        used_source.add(sc)
        used_target.add(tc)
        renames.append({"source_name": sc, "target_name_before": tc, "confidence": round(score, 2)})

    if rename_map:
        df_target = df_target.rename(columns=rename_map)
    return df_target, renames


def guess_key_columns(df_source, df_target):
    common = [col for col in df_source.columns if col in df_target.columns]
    if not common:
        # No shared columns — fall back to the first column of the source file
        return [df_source.columns[0]] if len(df_source.columns) else []
    lower_lookup = {col.lower(): col for col in common}

    # FIX: greatly expanded candidate list to cover real-world file naming
    # conventions beyond the original 6 hardcoded names.
    key_candidates = (
        # exact common names
        "id", "key", "record_id", "transaction_id", "customer_id", "account_id",
        "invoice_id", "order_id", "employee_id", "product_id", "ref_id", "case_id",
        "claim_id", "policy_id", "loan_id", "contract_id", "project_id", "ticket_id",
        # common abbreviations / short names
        "ref", "code", "no", "num", "number", "seq", "serial",
        "invoice_no", "invoice_number", "order_no", "order_number",
        "txn_id", "txn", "trx_id", "trx", "trans_id",
        "emp_id", "cust_id", "acct_id", "proj_id",
        # spaced / title-case variants people use in Excel headers
        "invoice no", "invoice number", "order no", "order number",
        "transaction id", "customer id", "account id", "employee id",
        "project name", "project id", "reference", "reference no",
        "reference number", "ref no", "ref number",
        "unique id", "unique key", "primary key",
    )
    for candidate in key_candidates:
        if candidate in lower_lookup:
            return [lower_lookup[candidate]]

    # FIX: if no known name matched, score every common column by uniqueness
    # ratio in the source file and pick the most unique one — a true key
    # column will have ~100% unique values, while a data column won't.
    best_col = None
    best_ratio = 0.0
    for col in common:
        col_series = df_source[col].astype(str).str.strip()
        non_empty = col_series[col_series != ""]
        if non_empty.empty:
            continue
        ratio = non_empty.nunique() / len(non_empty)
        if ratio > best_ratio:
            best_ratio = ratio
            best_col = col
    # Only trust the uniqueness heuristic if the best column is at least 90%
    # unique — otherwise fall back to the first common column as before.
    if best_col and best_ratio >= 0.9:
        return [best_col]
    return [common[0]]


DATA_TYPES = {"master", "transactional"}


def resolve_data_type(requested_type, df_source, df_target, key_columns):
    """Return the selected data type, or make a conservative automatic choice.

    Master data normally has one record per business key.  Transactional
    extracts commonly contain several lines for a key (invoice/order/account)
    and include amount/date/document fields.  The UI can always override this
    inference when a business has more specific knowledge of an extract.
    """
    requested_type = (requested_type or "auto").strip().lower()
    if requested_type in DATA_TYPES:
        return requested_type
    if requested_type not in ("", "auto"):
        raise ValueError("data_type must be 'auto', 'master', or 'transactional'.")

    common_columns = [str(c).lower() for c in df_source.columns if c in df_target.columns]
    transaction_words = ("transaction", "invoice", "order", "payment", "amount", "debit", "credit", "balance", "date", "journal", "document")
    has_transaction_shape = sum(any(word in col for word in transaction_words) for col in common_columns) >= 2
    has_transaction_identifier = any(
        word in col for col in common_columns
        for word in ("transaction", "invoice", "order", "payment", "journal", "document")
    )
    if not key_columns:
        return "transactional" if has_transaction_shape else "master"

    source_duplicate_ratio = df_source.duplicated(subset=key_columns, keep=False).mean() if len(df_source) else 0
    target_duplicate_ratio = df_target.duplicated(subset=key_columns, keep=False).mean() if len(df_target) else 0
    if has_transaction_shape and has_transaction_identifier:
        return "transactional"
    if max(source_duplicate_ratio, target_duplicate_ratio) >= 0.02:
        return "transactional"
    return "master"


def detect_date_column(df_source, df_target):
    lower_source = {col.lower(): col for col in df_source.columns}
    lower_target = {col.lower(): col for col in df_target.columns}
    common_lower = [k for k in lower_source if k in lower_target]

    for candidate in DATE_COLUMNS:
        key = candidate.lower()
        if key in common_lower:
            return lower_source[key]

    date_like = [k for k in common_lower if "date" in k]
    if date_like:
        date_like.sort(key=lambda k: (not k.endswith("date"), k))
        return lower_source[date_like[0]]

    best_col = None
    best_ratio = 0.0
    for key in common_lower:
        col = lower_source[key]
        series = df_source[col].dropna().astype(str).str.strip().head(50)
        series = series[series != ""]
        if series.empty:
            continue
        parsed = pd.to_datetime(series, errors="coerce")
        ratio = parsed.notna().mean()
        if ratio > 0.8 and ratio > best_ratio:
            best_ratio = ratio
            best_col = col
    return best_col


# Regex to guard against pd.to_datetime treating plain numbers as dates.
import re as _re
_PLAIN_NUMBER_RE = _re.compile(r'^[\$\-\+]?[\d,]+\.?\d*$')


def canonical_value(value):
    text = "" if pd.isna(value) else str(value)
    stripped = text.strip()
    if stripped == "":
        return ""
    number_text = stripped.replace(",", "")
    if number_text.startswith("$"):
        number_text = number_text[1:]
    try:
        return f"number:{Decimal(number_text).normalize()}"
    except InvalidOperation:
        pass
    # FIX: only attempt date parsing when the value actually looks like a date
    # (contains a separator like - / . or a month name). Plain numbers such as
    # "2024", "100", "12" were previously parsed as dates by pd.to_datetime,
    # causing false mismatches when the same number appeared as a numeric value
    # in one file and as a "date" in the other.
    if not _PLAIN_NUMBER_RE.match(stripped):
        parsed_date = pd.to_datetime(stripped, errors="coerce")
        if not pd.isna(parsed_date):
            return f"date:{parsed_date.date().isoformat()}"
    return f"text:{stripped.casefold()}"


def row_key_series(df, key_columns):
    return df[key_columns].astype(str).apply(
        lambda row: "||".join([cell.strip() for cell in row]), axis=1
    )


@functools.lru_cache(maxsize=16384)
def _date_key_cached(value_str):
    parsed = pd.to_datetime(value_str, errors="coerce")
    if pd.isna(parsed):
        return "Undated"
    return parsed.date().isoformat()


def date_key(value):
    value_str = "" if pd.isna(value) else str(value).strip()
    if not value_str:
        return "Undated"
    return _date_key_cached(value_str)


def records_with_key(df, indexes, keys):
    rows = []
    for idx in indexes:
        record = df.loc[idx].to_dict()
        record = {k: ("" if pd.isna(v) else str(v)) for k, v in record.items()}
        record["_reconciliation_key"] = keys.loc[idx]
        rows.append(record)
    return rows


def _key_text(row, key_columns):
    return " ".join(str(row.get(col, "")) for col in key_columns).strip()


def _diff_row_pair(source_row, target_row, compare_columns):
    row_mismatches = []
    row_formats = []
    for col in compare_columns:
        left_text = str(source_row.get(col, ""))
        right_text = str(target_row.get(col, ""))
        left_canonical = canonical_value(left_text)
        right_canonical = canonical_value(right_text)

        if left_text.strip() != right_text.strip() and left_canonical == right_canonical:
            row_formats.append({
                "column": col,
                "source_value": left_text,
                "target_value": right_text,
                "normalized_value": left_canonical.split(":", 1)[-1],
            })
        elif left_canonical != right_canonical:
            row_mismatches.append({
                "column": col,
                "source_value": left_text,
                "target_value": right_text,
            })
    return row_mismatches, row_formats


def transactional_difference_summary(df_source, df_target, key_columns):
    """Reconcile transaction lines without collapsing repeated business keys.

    A master-data comparison treats a duplicate key as an exception and keeps
    only one row.  That loses valid AP/AR lines when, for example, one invoice
    has multiple postings.  Here we first pair identical lines inside each key
    group, then pair the remaining lines in their file order so every source
    and target transaction is accounted for exactly once.
    """
    source_keys = row_key_series(df_source, key_columns)
    target_keys = row_key_series(df_target, key_columns)
    date_col = detect_date_column(df_source, df_target)
    compare_columns = [c for c in df_source.columns if c in df_target.columns and c not in key_columns]
    source_only_columns = [c for c in df_source.columns if c not in df_target.columns]
    target_only_columns = [c for c in df_target.columns if c not in df_source.columns]

    def clean_row(df, index):
        return {k: ("" if pd.isna(v) else str(v)) for k, v in df.loc[index].to_dict().items()}

    def fingerprint(row):
        return "||".join(canonical_value(row.get(c, "")) for c in compare_columns)

    source_groups, target_groups = {}, {}
    for index, key in source_keys.items(): source_groups.setdefault(key, []).append(index)
    for index, key in target_keys.items(): target_groups.setdefault(key, []).append(index)

    mismatch_rows, format_rows, full_rows = [], [], []
    missing_target, missing_source = [], []
    for key in sorted(set(source_groups) | set(target_groups)):
        source_indexes = source_groups.get(key, [])[:]
        target_indexes = target_groups.get(key, [])[:]
        # Match equal lines first. This makes re-ordered transaction exports clean.
        target_by_fingerprint = {}
        for index in target_indexes:
            target_by_fingerprint.setdefault(fingerprint(clean_row(df_target, index)), []).append(index)
        exact_pairs, remaining_source = [], []
        used_target = set()
        for index in source_indexes:
            candidates = target_by_fingerprint.get(fingerprint(clean_row(df_source, index)), [])
            target_index = next((candidate for candidate in candidates if candidate not in used_target), None)
            if target_index is None:
                remaining_source.append(index)
            else:
                used_target.add(target_index)
                exact_pairs.append((index, target_index))
        remaining_target = [index for index in target_indexes if index not in used_target]

        for source_index, target_index in exact_pairs + list(zip(remaining_source, remaining_target)):
            source_row, target_row = clean_row(df_source, source_index), clean_row(df_target, target_index)
            differences, formats = _diff_row_pair(source_row, target_row, compare_columns)
            key_data = {col: source_row.get(col, "") for col in key_columns}
            row_date = date_key(source_row.get(date_col, "")) if date_col else "Undated"
            if differences:
                mismatch_rows.append({"key": key_data, "date": row_date, "differences": differences,
                                      "changed_columns": [d["column"] for d in differences], "source_row": source_row, "target_row": target_row})
            if formats:
                format_rows.append({"key": key_data, "date": row_date, "differences": formats,
                                    "changed_columns": [d["column"] for d in formats], "source_row": source_row, "target_row": target_row})
            full_rows.append({"key": key_data, "status": "Updated" if differences else ("Format Only" if formats else "Matched"),
                              "changed_columns": [d["column"] for d in differences] + [d["column"] for d in formats],
                              "source_row": source_row, "target_row": target_row})

        for index in remaining_source[len(remaining_target):]:
            row = clean_row(df_source, index)
            row["_reconciliation_key"] = key
            missing_target.append(row)
            full_rows.append({"key": {col: row.get(col, "") for col in key_columns}, "status": "Deleted", "changed_columns": [], "source_row": {k: v for k, v in row.items() if k != "_reconciliation_key"}, "target_row": {}})
        for index in remaining_target[len(remaining_source):]:
            row = clean_row(df_target, index)
            row["_reconciliation_key"] = key
            missing_source.append(row)
            full_rows.append({"key": {col: row.get(col, "") for col in key_columns}, "status": "Added", "changed_columns": [], "source_row": {}, "target_row": {k: v for k, v in row.items() if k != "_reconciliation_key"}})

    # Detect amount column and compute invoice totals for transactional EDA
    amount_keywords = ("amount", "amt", "value", "total", "price", "sum", "debit", "credit")
    amount_col = next((c for c in df_source.columns if any(w in c.lower() for w in amount_keywords) and c in df_target.columns), None)
    invoice_summary = None
    if amount_col:
        def _to_float_series(s):
            return pd.to_numeric(s.astype(str).str.replace(r"[\$,]", "", regex=True), errors="coerce")
        src_total = float(_to_float_series(df_source[amount_col]).sum())
        tgt_total = float(_to_float_series(df_target[amount_col]).sum())
        invoice_summary = {
            "amount_column": amount_col,
            "source_invoice_total": round(src_total, 2),
            "target_invoice_total": round(tgt_total, 2),
            "invoice_difference": round(src_total - tgt_total, 2),
        }

    return {
        "source_record_count": int(len(df_source)), "target_record_count": int(len(df_target)),
        "date_column": date_col, "data_type": "transactional",
        "schema": {"source_columns": list(df_source.columns), "target_columns": list(df_target.columns), "source_only_columns": source_only_columns, "target_only_columns": target_only_columns},
        "missing_in_target": {"count": len(missing_target), "rows": missing_target},
        "missing_in_source": {"count": len(missing_source), "rows": missing_source},
        # Repeated keys are expected for transaction data, not duplicate defects.
        "duplicates_source": {"count": 0, "rows": []}, "duplicates_target": {"count": 0, "rows": []},
        "mismatches": {"count": len(mismatch_rows), "rows": mismatch_rows},
        "format_inconsistencies": {"count": len(format_rows), "rows": format_rows},
        "fuzzy_matches": {"count": 0, "rows": []}, "full_comparison": {"count": len(full_rows), "rows": full_rows},
    }


def difference_summary(df_source, df_target, key_columns, data_type="master"):
    if data_type == "transactional":
        return transactional_difference_summary(df_source, df_target, key_columns)
    source_keys = row_key_series(df_source, key_columns)
    target_keys = row_key_series(df_target, key_columns)
    date_col = detect_date_column(df_source, df_target)

    missing_in_target_idx = source_keys[~source_keys.isin(target_keys)].index.tolist()
    missing_in_source_idx = target_keys[~target_keys.isin(source_keys)].index.tolist()

    duplicates_source = df_source[df_source.duplicated(subset=key_columns, keep=False)]
    duplicates_target = df_target[df_target.duplicated(subset=key_columns, keep=False)]

    source_unique = df_source.assign(_reconciliation_key=source_keys).drop_duplicates(subset=key_columns, keep="first")
    target_unique = df_target.assign(_reconciliation_key=target_keys).drop_duplicates(subset=key_columns, keep="first")
    merged = source_unique.merge(target_unique, on=key_columns, how="inner", suffixes=("_src", "_tgt"))
    mismatch_rows = []
    format_rows = []
    full_comparison_rows = []
    source_only_columns = [c for c in df_source.columns if c not in df_target.columns]
    target_only_columns = [c for c in df_target.columns if c not in df_source.columns]
    compare_columns = [c for c in df_source.columns if c in df_target.columns and c not in key_columns]

    deleted_key_texts = {idx: _key_text(df_source.loc[idx], key_columns) for idx in missing_in_target_idx}
    added_key_texts = {idx: _key_text(df_target.loc[idx], key_columns) for idx in missing_in_source_idx}
    fuzzy_pairs, missing_in_target_idx, missing_in_source_idx = find_fuzzy_matches(
        deleted_key_texts, added_key_texts
    )

    fuzzy_rows = []
    for source_idx, target_idx, confidence in fuzzy_pairs:
        source_row_full = {k: ("" if pd.isna(v) else str(v)) for k, v in df_source.loc[source_idx].to_dict().items()}
        target_row_full = {k: ("" if pd.isna(v) else str(v)) for k, v in df_target.loc[target_idx].to_dict().items()}
        row_mismatches, row_formats = _diff_row_pair(source_row_full, target_row_full, compare_columns)
        changed_columns = [d["column"] for d in row_mismatches] + [d["column"] for d in row_formats]
        fuzzy_row = {
            "key_before": {col: source_row_full.get(col, "") for col in key_columns},
            "key_after": {col: target_row_full.get(col, "") for col in key_columns},
            "confidence": confidence,
            "changed_columns": changed_columns,
            "differences": row_mismatches,
            "format_differences": row_formats,
            "source_row": source_row_full,
            "target_row": target_row_full,
        }
        fuzzy_rows.append(fuzzy_row)
        full_comparison_rows.append({
            "key": fuzzy_row["key_after"],
            "status": "Renamed",
            "changed_columns": changed_columns,
            "source_row": source_row_full,
            "target_row": target_row_full,
            "match_confidence": confidence,
        })

    merged_records = merged.to_dict(orient="records")
    for row in merged_records:
        row_mismatches = []
        row_formats = []
        for col in compare_columns:
            left = row.get(f"{col}_src", "")
            right = row.get(f"{col}_tgt", "")
            left_text = "" if pd.isna(left) else str(left)
            right_text = "" if pd.isna(right) else str(right)
            left_canonical = canonical_value(left_text)
            right_canonical = canonical_value(right_text)

            if left_text.strip() != right_text.strip() and left_canonical == right_canonical:
                row_formats.append({
                    "column": col,
                    "source_value": left_text,
                    "target_value": right_text,
                    "normalized_value": left_canonical.split(":", 1)[-1],
                })
            elif left_canonical != right_canonical:
                row_mismatches.append({
                    "column": col,
                    "source_value": left_text,
                    "target_value": right_text,
                })

        source_row_full = {col: row.get(col, "") for col in key_columns}
        source_row_full.update({col: row.get(f"{col}_src", "") for col in compare_columns})
        source_row_full.update({col: row.get(col, "") for col in source_only_columns})

        target_row_full = {col: row.get(col, "") for col in key_columns}
        target_row_full.update({col: row.get(f"{col}_tgt", "") for col in compare_columns})
        target_row_full.update({col: row.get(col, "") for col in target_only_columns})

        row_date = date_key(row.get(f"{date_col}_src", row.get(date_col, ""))) if date_col else "Undated"

        if row_mismatches:
            mismatch_rows.append({
                "key": {col: row.get(col, "") for col in key_columns},
                "date": row_date,
                "differences": row_mismatches,
                "changed_columns": [d["column"] for d in row_mismatches],
                "source_row": source_row_full,
                "target_row": target_row_full,
            })
        if row_formats:
            format_rows.append({
                "key": {col: row.get(col, "") for col in key_columns},
                "date": row_date,
                "differences": row_formats,
                "changed_columns": [d["column"] for d in row_formats],
                "source_row": source_row_full,
                "target_row": target_row_full,
            })

        if row_mismatches:
            row_status = "Updated"
        elif row_formats:
            row_status = "Format Only"
        else:
            row_status = "Matched"
        full_comparison_rows.append({
            "key": {col: row.get(col, "") for col in key_columns},
            "status": row_status,
            "changed_columns": [d["column"] for d in row_mismatches] + [d["column"] for d in row_formats],
            "source_row": source_row_full,
            "target_row": target_row_full,
        })

    for entry in records_with_key(df_source, missing_in_target_idx, source_keys):
        entry = dict(entry)
        entry.pop("_reconciliation_key", None)
        full_comparison_rows.append({
            "key": {col: entry.get(col, "") for col in key_columns},
            "status": "Deleted", "changed_columns": [],
            "source_row": entry, "target_row": {},
        })

    for entry in records_with_key(df_target, missing_in_source_idx, target_keys):
        entry = dict(entry)
        entry.pop("_reconciliation_key", None)
        full_comparison_rows.append({
            "key": {col: entry.get(col, "") for col in key_columns},
            "status": "Added", "changed_columns": [],
            "source_row": {}, "target_row": entry,
        })

    return {
        "source_record_count": int(len(df_source)),
        "target_record_count": int(len(df_target)),
        "data_type": "master",
        "date_column": date_col,
        "schema": {
            "source_columns": list(df_source.columns),
            "target_columns": list(df_target.columns),
            "source_only_columns": source_only_columns,
            "target_only_columns": target_only_columns,
        },
        "missing_in_target": {"count": len(missing_in_target_idx), "rows": records_with_key(df_source, missing_in_target_idx, source_keys)},
        "missing_in_source": {"count": len(missing_in_source_idx), "rows": records_with_key(df_target, missing_in_source_idx, target_keys)},
        "duplicates_source": {"count": len(duplicates_source), "rows": duplicates_source.to_dict(orient="records")},
        "duplicates_target": {"count": len(duplicates_target), "rows": duplicates_target.to_dict(orient="records")},
        "mismatches": {"count": len(mismatch_rows), "rows": mismatch_rows},
        "format_inconsistencies": {"count": len(format_rows), "rows": format_rows},
        "fuzzy_matches": {"count": len(fuzzy_rows), "rows": fuzzy_rows},
        "full_comparison": {"count": len(full_comparison_rows), "rows": full_comparison_rows},
    }


def extract_day_summary(df_source, df_target, key_columns, diff_report):
    date_col = diff_report.get("date_column")
    if not date_col:
        return []
    source = df_source.copy()
    target = df_target.copy()
    source["_day"] = source[date_col].apply(date_key)
    target["_day"] = target[date_col].apply(date_key)
    all_days = sorted(set(source["_day"]).union(set(target["_day"])))
    
    from datetime import datetime
    today_str = datetime.now().date().isoformat()
    all_days = [day for day in all_days if day == "Undated" or day <= today_str]

    summary = []

    def count_rows(rows, day):
        return sum(1 for row in rows if date_key(row.get(date_col, "")) == day)

    def count_issue_rows(rows, day):
        return sum(1 for row in rows if row.get("date", "Undated") == day)

    for day in all_days:
        summary.append({
            "date": day,
            "source_records": int((source["_day"] == day).sum()),
            "target_records": int((target["_day"] == day).sum()),
            "missing_in_target": count_rows(diff_report["missing_in_target"]["rows"], day),
            "missing_in_source": count_rows(diff_report["missing_in_source"]["rows"], day),
            "duplicates_source": count_rows(diff_report["duplicates_source"]["rows"], day),
            "duplicates_target": count_rows(diff_report["duplicates_target"]["rows"], day),
            "mismatches": count_issue_rows(diff_report["mismatches"]["rows"], day),
            "format_inconsistencies": count_issue_rows(diff_report["format_inconsistencies"]["rows"], day),
        })
    return summary



# ── Health ────────────────────────────────────────────────────────────────────

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({"status": "ok"})


@app.route('/api/db/status', methods=['GET'])
def db_status():
    return jsonify({"connected": db.is_available()})


# ── Chat helpers ──────────────────────────────────────────────────────────────

def _top_mismatched_columns(diff_report, limit=5):
    counter = Counter()
    for bucket in ("mismatches", "format_inconsistencies", "fuzzy_matches"):
        for row in (diff_report.get(bucket) or {}).get("rows", []):
            for col in row.get("changed_columns") or []:
                counter[col] += 1
    return [{"column": col, "changes": n} for col, n in counter.most_common(limit)]


def _sample_keys(rows, key_columns, limit=8):
    samples = []
    for row in rows[:limit]:
        key = row.get("key") if isinstance(row, dict) and "key" in row else row
        if isinstance(key, dict):
            samples.append(", ".join(f"{k}={key.get(k, '')}" for k in (key_columns or key.keys())))
        else:
            samples.append(str(key))
    return samples


def build_dataset_chat_context(series_id, version=None, user_id=None):
    """Load reconciliation context for the AI assistant.

    When user_id is supplied (authenticated request) we verify ownership
    before returning anything — a user can never query another user's data.
    Returns (context_dict, error_message).
    """
    series = get_series(series_id)
    if not series:
        return None, "Selected dataset could not be found. It may have been deleted — please pick another dataset."

    # Ownership check: if we know who's asking, verify they own this series.
    # Admins bypass ownership checks entirely (see 3. ADMIN PERMISSIONS).
    if user_id is not None and not getattr(g, 'is_admin', False) and db.is_available():
        owner = db.get_series_owner(series_id)
        if owner is not None and owner != user_id:
            return None, "Access denied — this dataset belongs to another user."

    versions = series.get("versions", [])
    if not versions:
        return None, f"Dataset '{series.get('name', series_id)}' has no data yet."

    if version is None:
        diff_versions = [v for v in versions if v["version"] > 0]
        version_entry = diff_versions[-1] if diff_versions else versions[-1]
        version = version_entry["version"]
    else:
        version_entry = next((v for v in versions if v["version"] == version), None)
        if version_entry is None:
            return None, f"Version {version} was not found in dataset '{series['name']}'."

    context = {
        "dataset_name": series["name"],
        "version": version,
        "version_label": version_entry.get("label"),
        "row_count": version_entry.get("row_count"),
        "column_count": version_entry.get("column_count"),
        "key_columns": version_entry.get("key_columns") or [],
    }

    if version == 0 or not version_entry.get("diff_summary"):
        context["status"] = "baseline_only"
        context["note"] = (
            f"Only the baseline file (Version 0, '{version_entry.get('label')}') has been uploaded for "
            f"'{series['name']}' — {version_entry.get('row_count', 0)} rows, "
            f"{version_entry.get('column_count', 0)} columns. No reconciliation comparison has been run yet."
        )
        return context, None

    diff_summary = version_entry.get("diff_summary") or {}
    context["compared_against_version"] = diff_summary.get("compared_against_version")
    context["compared_against_label"] = diff_summary.get("compared_against_label")

    diff_report = load_series_diff_json(series_id, version)

    if diff_report is None:
        context["status"] = "summary_only"
        context["note"] = (
            "The full row-level reconciliation report file was not found on disk; "
            "only the summary counts below are available for this version."
        )
        context["stats"] = {
            "added": diff_summary.get("added", 0),
            "deleted": diff_summary.get("deleted", 0),
            "updated": diff_summary.get("updated", 0),
            "renamed": diff_summary.get("renamed", 0),
            "duplicates": diff_summary.get("duplicates", 0),
            "format_issues": diff_summary.get("format_issues", 0),
        }
        return context, None

    context["status"] = "full_report"
    context["source_record_count"] = diff_report.get("source_record_count")
    context["target_record_count"] = diff_report.get("target_record_count")
    context["stats"] = {
        "missing_in_target_deleted": diff_report.get("missing_in_target", {}).get("count", 0),
        "missing_in_source_added": diff_report.get("missing_in_source", {}).get("count", 0),
        "mismatches_updated_values": diff_report.get("mismatches", {}).get("count", 0),
        "fuzzy_renamed_matches": diff_report.get("fuzzy_matches", {}).get("count", 0),
        "duplicates_in_source": diff_report.get("duplicates_source", {}).get("count", 0),
        "duplicates_in_target": diff_report.get("duplicates_target", {}).get("count", 0),
        "format_inconsistencies": diff_report.get("format_inconsistencies", {}).get("count", 0),
    }
    context["top_mismatched_columns"] = _top_mismatched_columns(diff_report)

    key_cols = context["key_columns"]
    context["sample_missing_in_target_deleted"] = _sample_keys(
        diff_report.get("missing_in_target", {}).get("rows", []), key_cols)
    context["sample_missing_in_source_added"] = _sample_keys(
        diff_report.get("missing_in_source", {}).get("rows", []), key_cols)
    context["sample_duplicate_keys_source"] = _sample_keys(
        diff_report.get("duplicates_source", {}).get("rows", []), key_cols)
    context["sample_duplicate_keys_target"] = _sample_keys(
        diff_report.get("duplicates_target", {}).get("rows", []), key_cols)

    day_summary = diff_report.get("day_summary") or []
    if day_summary:
        context["day_summary"] = day_summary

    insights = diff_report.get("insights") or {}
    if insights.get("narrative"):
        context["narrative_summary"] = insights["narrative"]
    if insights.get("churn_percent") is not None:
        context["churn_percent"] = insights["churn_percent"]
        context["churn_label"] = insights.get("churn_label")

    return context, None


def build_reconciliation_prompt(message, context, history):
    system_instructions = (
        "You are an AI Data Reconciliation Assistant.\n"
        "Answer ONLY using the supplied reconciliation context below. Never invent, guess, or "
        "estimate values that are not present in the context. If the information needed to answer "
        "is not available in the context, say plainly that it is unavailable.\n"
        "Explain reconciliation statistics in simple, plain language a non-technical business user "
        "can understand. Provide concise but informative answers, and reference concrete numbers "
        "from the context when relevant.\n"
    )

    prompt = system_instructions

    if context.get("status") == "baseline_only":
        prompt += f"\n<reconciliation context>\n{context['note']}\n</reconciliation context>\n"
    else:
        prompt += f"\n<reconciliation context>\n{json.dumps(context, default=str, indent=2)}\n</reconciliation context>\n"

    if history:
        prompt += "\nPrior conversation:\n"
        for turn in history[-10:]:
            role = 'User' if turn.get('role') == 'user' else 'Assistant'
            content = (turn.get('content') or '').strip()
            if content:
                prompt += f"{role}: {content}\n"

    prompt += f"\nUser: {message}"
    return prompt


# ── Chat API ──────────────────────────────────────────────────────────────────

@app.route('/api/chat', methods=['POST'])
@optional_auth
def chat():
    """Dataset-aware chatbot endpoint.

    Accepts:
      message     (str, required)
      series_id   (str, required) — which dataset to answer from
      version     (int, optional) — specific version; defaults to latest diff
      history     (list, optional) — prior turns for context
      provider    (str, optional) — 'groq', 'ollama', or '' (empty = auto;
                   Groq first, then automatic fallback to Ollama)
      model       (str, optional) — override the provider's default model
                   (GROQ_MODEL for groq/auto, OLLAMA_MODEL for ollama)

    When a valid JWT is present the user_id is used to enforce ownership:
    the assistant will refuse to answer questions about another user's dataset.
    Unauthenticated requests (dev/legacy) still work but skip the ownership check.
    """
    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    series_id = data.get('series_id')
    version = data.get('version')
    history = data.get('history') or []
    provider = (data.get('provider') or '').strip().lower()
    model = data.get('model')

    if not message:
        return jsonify({"error": "message is required."}), 400

    if not series_id:
        return jsonify({"error": "Please select a dataset before asking questions."}), 400

    if provider not in ('', 'auto', 'groq', 'ollama'):
        return jsonify({"error": "provider must be 'groq', 'ollama', or '' (auto)."}), 400

    if version is not None:
        try:
            version = int(version)
        except (TypeError, ValueError):
            return jsonify({"error": "version must be a number."}), 400

    # g.current_user_id is set by @optional_auth — None when unauthenticated.
    user_id = getattr(g, 'current_user_id', None)

    from kafka_service import KAFKA_ENABLED
    if KAFKA_ENABLED:
        import uuid
        job_id = f"chat_{uuid.uuid4().hex[:16]}"
        payload_params = {
            "job_id": job_id,
            "job_type": "CHAT_RESPONSE",
            "user_id": user_id,
            "message": message,
            "series_id": series_id,
            "version": version,
            "history": history,
            "provider": provider,
            "model": model,
        }
        db.create_recon_job(
            job_id=job_id,
            user_id=user_id,
            job_type="CHAT_RESPONSE",
            status="QUEUED_KAFKA",
            payload_params=payload_params
        )
        publish_recon_job(job_id, payload_params)
        return jsonify({"job_id": job_id, "async": True}), 202

    context, error = build_dataset_chat_context(series_id, version, user_id=user_id)
    if error:
        return jsonify({"error": error}), 404

    prompt = build_reconciliation_prompt(message, context, history)

    note = None
    if provider == 'groq':
        # Explicit Groq. If the chosen model is rate-limited (free tier's
        # per-minute/per-day token caps) or otherwise fails, fall back to
        # Ollama so the chat still answers, and tell the user about it.
        try:
            response_text = groq_generate(prompt, model=model)
        except GroqError as groq_exc:
            # Never forward the Groq model id to Ollama — use Ollama's default.
            try:
                response_text = generate_response(prompt)
            except OllamaError as ollama_exc:
                return jsonify({"error": f"Groq: {groq_exc}; Ollama: {ollama_exc}"}), 503
            note = (
                f"Groq model '{model or 'default'}' was unavailable ({groq_exc}). "
                "Answered by Ollama instead."
            )
    elif provider == 'ollama':
        try:
            response_text = generate_response(prompt, model=model)
        except OllamaError as exc:
            return jsonify({"error": str(exc)}), 503
    else:
        # Auto: Groq is primary, Ollama is the automatic fallback.
        try:
            response_text = groq_generate(prompt, model=model)
        except GroqError as groq_exc:
            try:
                response_text = generate_response(prompt)
            except OllamaError as ollama_exc:
                return jsonify({"error": f"Groq: {groq_exc}; Ollama: {ollama_exc}"}), 503

    payload = {"response": response_text, "context": context}
    if note:
        payload["note"] = note
    return jsonify(payload)


# ── One-off reconcile (no series) ────────────────────────────────────────────

@app.route('/api/reconcile', methods=['POST'])
@optional_auth
def reconcile():
    if 'source_file' not in request.files or 'target_file' not in request.files:
        return jsonify({"error": "Please upload both source_file and target_file."}), 400

    source_file = request.files['source_file']
    target_file = request.files['target_file']

    if source_file.filename == '' or target_file.filename == '':
        return jsonify({"error": "Both files must have a filename."}), 400

    if not allowed_file(source_file.filename) or not allowed_file(target_file.filename):
        return jsonify({"error": "Allowed file types: csv, xls, xlsx."}), 400

    try:
        df_source = normalize_dataframe(read_dataframe(source_file))
        df_target = normalize_dataframe(read_dataframe(target_file))
    except Exception as exc:
        return jsonify({"error": f"Could not read files: {str(exc)}"}), 400

    mapping_mode = request.form.get('mapping_mode', 'HEADER_COLUMN').strip()
    row_mappings_raw = request.form.get('row_mappings')

    # ── MODE 2: ROW-TO-ROW MAPPING WITH INDEXING ────────────────────────────
    if mapping_mode == 'ROW_INDEX' or row_mappings_raw:
        try:
            row_mappings = json.loads(row_mappings_raw) if row_mappings_raw else []
        except Exception:
            return jsonify({"error": "Invalid row_mappings JSON payload."}), 400

        if not row_mappings:
            return jsonify({"error": "No row pairs were selected for Row-to-Row Mapping."}), 400

        try:
            tolerance = float(request.form.get('tolerance', 0.01))
            row_results = reconcile_by_row_indexing(df_source, df_target, row_mappings, tolerance=tolerance)
        except Exception as exc:
            return jsonify({"error": f"Row reconciliation failed: {str(exc)}"}), 400

        user_id = getattr(g, 'current_user_id', None)
        source_metadata = store_file(source_file.filename, df_source, "source", user_id=user_id)
        target_metadata = store_file(target_file.filename, df_target, "target", user_id=user_id)

        # Store session in DB
        session_id = request.form.get('session_id') or f"sess_{os.urandom(8).hex()}"
        db.save_mapping_session(
            session_id=session_id,
            user_id=user_id,
            source_dataset_id=source_file.filename,
            target_dataset_id=target_file.filename,
            mapping_mode="ROW_INDEX",
            row_mappings=row_mappings,
        )

        return jsonify({
            "mapping_mode": "ROW_INDEX",
            "session_id": session_id,
            "results": row_results,
            "summary": row_results["summary"],
            "stored_files": [source_metadata, target_metadata],
        })

    # ── MODE 1: HEADER / COLUMN MAPPING (NO INDEXING) ────────────────────────
    manual_mapping = _parse_schema_mapping_form(request.form)
    df_target, _manual_renames = apply_manual_schema_mapping(df_target, manual_mapping)


    # GENERIC fix (works for any op-co, not just one): rename Target columns
    # that are the same field as a Source column but styled differently
    # (e.g. 'PartyNumber' vs 'PARTY_NUMBER') onto Source's naming, so exact
    # -name key detection below actually finds them as shared.
    df_target, _col_alignments = align_equivalent_columns(df_source, df_target)

    # Third tier: anything still unmatched (and not explicitly skipped by
    # the user) gets a best-effort fuzzy/substring match -- e.g.
    # 'Address1_original' -> 'ADDRESS1' -- instead of staying "Ignored"
    # just because it didn't normalise to an exact match.
    df_target, _fuzzy_renames = fuzzy_align_remaining_columns(
        df_source, df_target, excluded_source_cols=explicitly_ignored_source_columns(manual_mapping)
    )

    key_columns = request.form.get('key_columns', '').strip()
    if key_columns:
        key_columns = [c.strip() for c in key_columns.split(',') if c.strip()]
    else:
        key_columns = guess_key_columns(df_source, df_target)

    if not key_columns:
        return jsonify({"error": "No key columns found. Provide key_columns or ensure common column names exist."}), 400

    for col in key_columns:
        if col not in df_source.columns and col not in df_target.columns:
            return jsonify({"error": f"Key column '{col}' was not found in either file."}), 400
        if col not in df_source.columns:
            df_source[col] = ""
        if col not in df_target.columns:
            df_target[col] = ""

    try:
        data_type = resolve_data_type(request.form.get("data_type"), df_source, df_target, key_columns)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    diff_report = difference_summary(df_source, df_target, key_columns, data_type)
    day_summary = extract_day_summary(df_source, df_target, key_columns, diff_report)
    insights = generate_plain_english_summary(diff_report, day_summary, key_columns, "Source", "Target")
    diff_report["insights"] = insights
    diff_report["schema_mapping"] = manual_mapping

    user_id = getattr(g, 'current_user_id', None)
    source_metadata = store_file(source_file.filename, df_source, "source", user_id=user_id)
    target_metadata = store_file(target_file.filename, df_target, "target", user_id=user_id)
    amt_src = request.form.get('amount_source_col') or None
    amt_tgt = request.form.get('amount_target_col') or None
    report_meta = store_report(diff_report, source_metadata, target_metadata, key_columns, day_summary, user_id=user_id,
                               amount_source_col=amt_src, amount_target_col=amt_tgt)

    return jsonify({
        "key_columns": key_columns,
        "data_type": data_type,
        "report": diff_report,
        "day_summary": day_summary,
        "insights": insights,
        "stored_files": [source_metadata, target_metadata],
        "report_meta": report_meta,
    })


# ── Stored files ──────────────────────────────────────────────────────────────

@app.route('/api/stored-files', methods=['GET'])
@optional_auth
def stored_files():
    user_id = getattr(g, 'current_user_id', None)
    is_admin = getattr(g, 'is_admin', False)
    files = list_files(user_id=user_id, is_admin=is_admin)
    return jsonify({"files": files})


@app.route('/api/stored-files', methods=['DELETE'])
@require_auth
def stored_files_delete_all():
    count = delete_all_files(user_id=g.current_user_id, is_admin=getattr(g, 'is_admin', False))
    return jsonify({"deleted": True, "count": count})


@app.route('/api/file-chunks/<file_id>', methods=['GET'])
@optional_auth
def file_chunks(file_id):
    data = get_file_chunks(file_id)
    if data is None:
        return jsonify({"error": "File not found."}), 404
    user_id = getattr(g, 'current_user_id', None)
    if user_id is not None and not getattr(g, 'is_admin', False):
        owner = data.get('metadata', {}).get('user_id')
        if owner is not None and owner != user_id:
            return jsonify({"error": "Access denied."}), 403
    return jsonify(data)


@app.route('/api/stored-files/<file_id>', methods=['DELETE'])
@require_auth
def stored_file_delete(file_id):
    if not getattr(g, 'is_admin', False):
        data = get_file_chunks(file_id)
        if data is not None:
            owner = data.get('metadata', {}).get('user_id')
            if owner is not None and owner != g.current_user_id:
                return jsonify({"error": "Access denied."}), 403
    deleted = delete_file(file_id)
    if not deleted:
        return jsonify({"error": "File not found."}), 404
    return jsonify({"deleted": True, "file_id": file_id})


@app.route('/api/preview-columns', methods=['POST'])
@app.route('/api/parse-columns', methods=['POST'])
@optional_auth
def preview_columns():
    """Read an uploaded file and return its column names + auto-detected key suggestion.
    Used by the frontend to show a column picker dropdown instead of a free-text input."""
    if 'file' not in request.files:
        return jsonify({"error": "Please upload a file as 'file'."}), 400
    uploaded = request.files['file']
    if uploaded.filename == '' or not allowed_file(uploaded.filename):
        return jsonify({"error": "Allowed file types: csv, xls, xlsx."}), 400
    try:
        df = normalize_dataframe(read_dataframe(uploaded))
    except Exception as exc:
        return jsonify({"error": f"Could not read file: {str(exc)}"}), 400
    columns = list(df.columns)
    # Reuse the same auto-detection logic so the suggestion matches what reconciliation would pick
    suggested = guess_key_columns(df, df)  # pass df twice — we only need column scoring, not cross-file matching
    return jsonify({"columns": columns, "suggested_key": suggested[0] if suggested else None})


# ── Dynamic Mapping Engine API Endpoints ─────────────────────────────────────

@app.route('/api/mapping/analyze-schema', methods=['POST'])
@optional_auth
def analyze_schema():
    """Generates dynamic 8-signal schema analysis, Hungarian optimal assignments, and key candidate scores."""
    if 'source_file' not in request.files or 'target_file' not in request.files:
        return jsonify({"error": "Both source_file and target_file are required for schema analysis."}), 400

    s_file = request.files['source_file']
    t_file = request.files['target_file']

    if s_file.filename == '' or t_file.filename == '':
        return jsonify({"error": "Both uploaded files must have valid filenames."}), 400

    try:
        df_source = normalize_dataframe(read_dataframe(s_file))
        df_target = normalize_dataframe(read_dataframe(t_file))
    except Exception as exc:
        return jsonify({"error": f"Could not parse uploaded files for schema analysis: {str(exc)}"}), 400

    analysis = generate_schema_mapping_analysis(df_source, df_target)
    return jsonify(analysis)


@app.route('/api/mapping/row-preview', methods=['POST'])
@optional_auth
def row_preview():
    """Returns server-side paginated previews, index search, and value search for Row-to-Row Mapping mode."""
    if 'file' not in request.files:
        return jsonify({"error": "Please upload a file as 'file'."}), 400

    uploaded = request.files['file']
    prefix = request.form.get('prefix', 'SRC').upper()
    try:
        page = int(request.form.get('page', 1))
        page_size = int(request.form.get('page_size', 20))
    except ValueError:
        page, page_size = 1, 20

    search_query = request.form.get('search', '').strip()

    try:
        df = normalize_dataframe(read_dataframe(uploaded))
    except Exception as exc:
        return jsonify({"error": f"Could not read file for row preview: {str(exc)}"}), 400

    preview_data = get_row_previews(df, prefix=prefix, page=page, page_size=page_size, search_query=search_query)
    return jsonify(preview_data)


@app.route('/api/mapping/column-sum', methods=['POST'])
@optional_auth
def column_sum():
    """Returns the sum of all numeric values in the specified column of an uploaded file."""
    if 'file' not in request.files:
        return jsonify({"error": "Please upload a file as 'file'."}), 400

    col_name = request.form.get('column', '').strip()
    if not col_name:
        return jsonify({"error": "Please specify a 'column' parameter."}), 400

    try:
        df = normalize_dataframe(read_dataframe(request.files['file']))
    except Exception as exc:
        return jsonify({"error": f"Could not read file: {str(exc)}"}), 400

    if col_name not in df.columns:
        return jsonify({"error": f"Column '{col_name}' not found in the file."}), 400

    import re as _re
    def _parse_amount(val):
        try:
            return float(_re.sub(r'[,$ ]', '', str(val)))
        except (ValueError, TypeError):
            return 0.0

    total = sum(_parse_amount(v) for v in df[col_name])
    return jsonify({"column": col_name, "sum": round(total, 2), "row_count": len(df)})


@app.route('/api/mapping/auto-match-rows', methods=['POST'])
@optional_auth
def auto_match_rows():
    """Automates row lookup matching (Source entity_name_ora = Target PARTY_NAME, tie-breaker city_ora = CITY)."""
    if 'source_file' not in request.files or 'target_file' not in request.files:
        return jsonify({"error": "Both source_file and target_file are required for row auto-matching."}), 400

    s_file = request.files['source_file']
    t_file = request.files['target_file']

    src_name_col = request.form.get('src_name_col', '').strip()
    tgt_name_col = request.form.get('tgt_name_col', '').strip()
    src_city_col = request.form.get('src_city_col', '').strip()
    tgt_city_col = request.form.get('tgt_city_col', '').strip()
    src_state_col = request.form.get('src_state_col', '').strip()
    tgt_num_col = request.form.get('tgt_num_col', '').strip()

    try:
        df_source = normalize_dataframe(read_dataframe(s_file))
        df_target = normalize_dataframe(read_dataframe(t_file))
    except Exception as exc:
        return jsonify({"error": f"Could not parse uploaded files: {str(exc)}"}), 400

    result = auto_match_rows_by_keys(
        df_source,
        df_target,
        src_name_col=src_name_col,
        tgt_name_col=tgt_name_col,
        src_city_col=src_city_col,
        tgt_city_col=tgt_city_col,
        src_state_col=src_state_col,
        tgt_num_col=tgt_num_col,
    )
    return jsonify(result)


@app.route('/api/mapping/save', methods=['POST'])
@optional_auth
def save_mapping():
    """Persists a confirmed mapping configuration (Header Column mode or Row Index mode) to database."""
    payload = request.get_json(silent=True) or {}
    session_id = payload.get('session_id') or f"sess_{os.urandom(8).hex()}"
    mapping_mode = payload.get('mapping_mode')

    if mapping_mode not in ('HEADER_COLUMN', 'ROW_INDEX'):
        return jsonify({"error": "mapping_mode must be HEADER_COLUMN or ROW_INDEX."}), 400

    user_id = getattr(g, 'current_user_id', None)
    res = db.save_mapping_session(
        session_id=session_id,
        user_id=user_id,
        source_dataset_id=payload.get('source_dataset_id'),
        target_dataset_id=payload.get('target_dataset_id'),
        mapping_mode=mapping_mode,
        header_mappings=payload.get('header_mappings'),
        row_mappings=payload.get('row_mappings'),
    )
    publish_audit_event("MAPPING_CONFIG_SAVED", {"session_id": session_id, "mapping_mode": mapping_mode}, user_id=user_id)
    return jsonify(res)


@app.route('/api/mapping/resume/<session_id>', methods=['GET'])
@optional_auth
def resume_mapping(session_id):
    """Retrieves a previously saved mapping configuration."""
    session = db.get_mapping_session(session_id)
    if not session:
        return jsonify({"error": f"Mapping session '{session_id}' not found."}), 404
    return jsonify(session)


@app.route('/api/stored-files/upload', methods=['POST'])
@optional_auth
def stored_file_upload():
    if 'file' not in request.files:
        return jsonify({"error": "Please upload a file as 'file'."}), 400
    uploaded = request.files['file']
    if uploaded.filename == '' or not allowed_file(uploaded.filename):
        return jsonify({"error": "Allowed file types: csv, xls, xlsx."}), 400
    try:
        df = normalize_dataframe(read_dataframe(uploaded))
    except Exception as exc:
        return jsonify({"error": f"Could not read file: {str(exc)}"}), 400
    meta = store_file(uploaded.filename, df, "uploaded", user_id=getattr(g, 'current_user_id', None))
    return jsonify({"file": meta}), 201


@app.route('/api/stored-files/<file_id>/preview', methods=['GET'])
@optional_auth
def stored_file_preview(file_id):
    limit = int(request.args.get('limit', 200))
    data = get_file_chunks(file_id)
    if data is None:
        return jsonify({"error": "File not found."}), 404
    rows = []
    for chunk in (data.get('chunks') or []):
        text = chunk.get('text', '')
        record = {}
        for line in text.strip().split('\n'):
            if ': ' in line:
                col, val = line.split(': ', 1)
                record[col.strip()] = val.strip()
        if record:
            rows.append(record)
    return jsonify({
        "file_id": file_id,
        "filename": data['metadata']['filename'],
        "columns": list(rows[0].keys()) if rows else [],
        "rows": rows[:limit],
        "total": len(rows),
    })


# ── Reports ────────────────────────────────────────────────────────────────────

def _series_owner_lookup():
    """Returns a callable(series_id) -> user_id|None backed by Postgres,
    or None if Postgres isn't reachable (callers then fall back to the
    local series.json file's user_id field)."""
    if db.is_available():
        return db.get_series_owner
    return None


@app.route('/api/reports', methods=['GET'])
@optional_auth
def reports_list():
    user_id = getattr(g, 'current_user_id', None)
    is_admin = getattr(g, 'is_admin', False)
    return jsonify({"reports": list_reports(user_id=user_id, is_admin=is_admin, db_owner_lookup=_series_owner_lookup())})


@app.route('/api/reports', methods=['DELETE'])
@require_auth
def reports_delete_all():
    count = delete_all_reports(user_id=g.current_user_id, is_admin=getattr(g, 'is_admin', False), db_owner_lookup=_series_owner_lookup())
    return jsonify({"deleted": True, "count": count})


@app.route('/api/reports/<report_name>', methods=['GET'])
@optional_auth
def report_download(report_name):
    safe_name = os.path.basename(report_name)
    path = os.path.join(os.path.dirname(__file__), 'vector_store', 'reports', safe_name)
    if not os.path.exists(path):
        return jsonify({"error": "Report not found."}), 404
    user_id = getattr(g, 'current_user_id', None)
    if user_id is not None and not getattr(g, 'is_admin', False):
        owner = get_report_owner(safe_name, db_owner_lookup=_series_owner_lookup())
        if owner != user_id:
            return jsonify({"error": "Access denied."}), 403
    return send_file(path, as_attachment=True)


@app.route('/api/reports/<report_name>', methods=['DELETE'])
@require_auth
def report_delete(report_name):
    safe_name = os.path.basename(report_name)
    path = os.path.join(os.path.dirname(__file__), 'vector_store', 'reports', safe_name)
    if not os.path.exists(path):
        return jsonify({"error": "Report not found."}), 404
    if not getattr(g, 'is_admin', False):
        owner = get_report_owner(safe_name, db_owner_lookup=_series_owner_lookup())
        if owner != g.current_user_id:
            return jsonify({"error": "Access denied."}), 403
    os.remove(path)
    return jsonify({"deleted": True, "report_file": safe_name})


# ── Datasets endpoint (user-scoped view of series) ────────────────────────────

@app.route('/api/datasets', methods=['GET'])
@require_auth
def datasets_list():
    """Return all datasets (series) that belong to the authenticated user.
    Falls back to the full series list when Postgres is unavailable so the
    UI never breaks in offline/dev mode."""
    user_id = g.current_user_id

    if db.is_available():
        all_series = list_series()
        if getattr(g, 'is_admin', False):
            # Admin dashboard: system-wide data instead of one user's.
            user_series = all_series
        else:
            owned_ids = set(db.list_series_for_user(user_id))
            user_series = [s for s in all_series if s["series_id"] in owned_ids]
    else:
        # Postgres down: fall back to storage.py's file-based list filtered
        # by the user_id stored in series metadata (if present).
        user_series = list_series_for_user(user_id)

    return jsonify({"datasets": user_series})


# ── Series (comparison chains) ────────────────────────────────────────────────

@app.route('/api/series', methods=['POST'])
@optional_auth
def series_create():
    """Register a new version-chain series. Automatically creates a dataset
    record with the same name as the uploaded file (Feature 1).
    The series is linked to the authenticated user when a JWT is present."""
    if 'file' not in request.files:
        return jsonify({"error": "Please upload a source file as 'file'."}), 400

    source_file = request.files['file']
    if source_file.filename == '' or not allowed_file(source_file.filename):
        return jsonify({"error": "Allowed file types: csv, xls, xlsx."}), 400

    try:
        df_source = normalize_dataframe(read_dataframe(source_file))
    except Exception as exc:
        return jsonify({"error": f"Could not read file: {str(exc)}"}), 400

    # Dataset name defaults to filename without extension (Feature 1)
    raw_name = request.form.get('name', '').strip()
    if not raw_name:
        raw_name = source_file.filename.rsplit('.', 1)[0]

    user_id = getattr(g, 'current_user_id', None)

    # Create the file-based series record (storage.py)
    requested_data_type = request.form.get("data_type", "auto")
    try:
        data_type = resolve_data_type(requested_data_type, df_source, df_source, guess_key_columns(df_source, df_source))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    series = create_series(raw_name, source_file.filename, df_source, user_id=user_id, data_type=data_type)

    # Mirror into Postgres — series metadata + dataset record
    db.upsert_series_metadata(series["series_id"], series["name"], user_id=user_id)
    db.upsert_dataset(
        dataset_id=series["series_id"],
        dataset_name=series["name"],
        original_file_name=source_file.filename,
        user_id=user_id,
        record_count=int(len(df_source)),
        file_type=source_file.filename.rsplit('.', 1)[-1].lower(),
        column_names=list(df_source.columns),
    )

    return jsonify({"series": series}), 201


@app.route('/api/series', methods=['GET'])
@optional_auth
def series_list():
    """List series. When authenticated, returns only the user's own series."""
    user_id = getattr(g, 'current_user_id', None)
    if user_id is not None:
        if db.is_available():
            all_series = list_series()
            if getattr(g, 'is_admin', False):
                # Admin dashboard: every series across every user.
                return jsonify({"series": all_series})
            owned_ids = set(db.list_series_for_user(user_id))
            return jsonify({"series": [s for s in all_series if s["series_id"] in owned_ids]})
        else:
            return jsonify({"series": list_series_for_user(user_id)})
    return jsonify({"series": list_series()})


@app.route('/api/series/<series_id>', methods=['GET'])
@optional_auth
def series_detail(series_id):
    user_id = getattr(g, 'current_user_id', None)
    series = get_series(series_id)
    if not series:
        return jsonify({"error": "Series not found."}), 404

    # Ownership guard — admins bypass.
    if user_id is not None and not getattr(g, 'is_admin', False) and db.is_available():
        owner = db.get_series_owner(series_id)
        if owner is not None and owner != user_id:
            return jsonify({"error": "Access denied."}), 403

    timeline = [
        {
            "version": v["version"],
            "label": v["label"],
            "uploaded_at": v["uploaded_at"],
            "row_count": v["row_count"],
            "added": (v.get("diff_summary") or {}).get("added", 0),
            "deleted": (v.get("diff_summary") or {}).get("deleted", 0),
            "updated": (v.get("diff_summary") or {}).get("updated", 0),
            "renamed": (v.get("diff_summary") or {}).get("renamed", 0),
            "format_issues": (v.get("diff_summary") or {}).get("format_issues", 0),
        }
        for v in series["versions"]
    ]
    return jsonify({"series": series, "timeline": timeline})


@app.route('/api/series/<series_id>', methods=['DELETE'])
@require_auth
def series_delete(series_id):
    user_id = g.current_user_id
    if not getattr(g, 'is_admin', False) and db.is_available():
        owner = db.get_series_owner(series_id)
        if owner is not None and owner != user_id:
            return jsonify({"error": "Access denied."}), 403

    deleted = delete_series(series_id)
    if not deleted:
        return jsonify({"error": "Series not found."}), 404
    db.delete_series_from_db(series_id)
    return jsonify({"deleted": True, "series_id": series_id})


@app.route('/api/series', methods=['DELETE'])
@require_auth
def series_delete_all():
    # NOTE: intentionally NOT admin-bypassed — this is a bulk delete, and
    # letting an admin session accidentally wipe every user's data via a
    # "delete all" call meant for their own account is a destructive
    # footgun. Admins can still delete any individual series via
    # /api/series/<id> DELETE (that ownership check is bypassed above).
    user_id = g.current_user_id
    if db.is_available():
        owned_ids = set(db.list_series_for_user(user_id))
        all_series = list_series()
        series_to_delete = [s["series_id"] for s in all_series if s["series_id"] in owned_ids]
    else:
        series_to_delete = [s["series_id"] for s in list_series_for_user(user_id)]

    count = 0
    for sid in series_to_delete:
        if delete_series(sid):
            db.delete_series_from_db(sid)
            count += 1
    return jsonify({"deleted": True, "count": count})


# ── Series versions ───────────────────────────────────────────────────────────

@app.route('/api/series/<series_id>/versions', methods=['POST'])
@optional_auth
def series_add_version(series_id):
    """Upload the next day's file, reconcile it against the previous version."""
    user_id = getattr(g, 'current_user_id', None)

    series = get_series(series_id)
    if not series:
        return jsonify({"error": "Series not found."}), 404

    # Ownership check — admins bypass.
    if user_id is not None and not getattr(g, 'is_admin', False) and db.is_available():
        owner = db.get_series_owner(series_id)
        if owner is not None and owner != user_id:
            return jsonify({"error": "Access denied."}), 403

    if 'file' not in request.files:
        return jsonify({"error": "Please upload the new day's file as 'file'."}), 400

    new_file = request.files['file']
    if new_file.filename == '' or not allowed_file(new_file.filename):
        return jsonify({"error": "Allowed file types: csv, xls, xlsx."}), 400

    try:
        df_new = normalize_dataframe(read_dataframe(new_file))
    except Exception as exc:
        return jsonify({"error": f"Could not read file: {str(exc)}"}), 400

    prev_version_entry = series["versions"][-1]
    prev_version = prev_version_entry["version"]
    df_prev = load_version_dataframe(series_id, prev_version)
    if df_prev is None:
        return jsonify({"error": "Previous version data could not be loaded."}), 500

    # Manual mapping from the Schema Mapping modal (if the user configured
    # one) always wins over the automatic heuristic below, for whichever
    # columns it covers.
    manual_mapping = _parse_schema_mapping_form(request.form)
    df_new, _manual_renames = apply_manual_schema_mapping(df_new, manual_mapping)

    # GENERIC fix (works for any op-co, not just one): rename the new file's
    # columns that are the same field as a previous-version column but
    # styled differently (e.g. 'PartyNumber' vs 'PARTY_NUMBER') onto the
    # previous version's naming, so exact-name key detection below finds
    # them as shared instead of treating them as unrelated columns.
    df_new, _col_alignments = align_equivalent_columns(df_prev, df_new)

    # Third tier: anything still unmatched (and not explicitly skipped by
    # the user) gets a best-effort fuzzy/substring match -- e.g.
    # 'Address1_original' -> 'ADDRESS1' -- instead of staying "Ignored"
    # just because it didn't normalise to an exact match.
    df_new, _fuzzy_renames = fuzzy_align_remaining_columns(
        df_prev, df_new, excluded_source_cols=explicitly_ignored_source_columns(manual_mapping)
    )

    key_columns = request.form.get('key_columns', '').strip()
    if key_columns:
        key_columns = [c.strip() for c in key_columns.split(',') if c.strip()]
    else:
        key_columns = prev_version_entry.get("key_columns") or guess_key_columns(df_prev, df_new)

    if not key_columns:
        return jsonify({"error": "No key columns found. Provide key_columns or ensure common column names exist."}), 400

    for col in key_columns:
        if col not in df_new.columns and col not in df_prev.columns:
            return jsonify({"error": f"Key column '{col}' was not found in either file."}), 400
        if col not in df_new.columns:
            df_new[col] = ""
        if col not in df_prev.columns:
            df_prev[col] = ""

    try:
        data_type = resolve_data_type(
            request.form.get("data_type") or series.get("data_type", "auto"), df_prev, df_new, key_columns
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    diff_report = difference_summary(df_prev, df_new, key_columns, data_type)
    day_summary = extract_day_summary(df_prev, df_new, key_columns, diff_report)

    next_version = prev_version + 1
    label = request.form.get('label', '').strip() or f"Day {next_version}"

    insights = generate_plain_english_summary(
        diff_report, day_summary, key_columns, prev_version_entry["label"], label
    )

    diff_summary_meta = {
        "data_type": data_type,
        "added": diff_report["missing_in_source"]["count"],
        "deleted": diff_report["missing_in_target"]["count"],
        "duplicates": diff_report["duplicates_source"]["count"] + diff_report["duplicates_target"]["count"],
        "updated": diff_report["mismatches"]["count"],
        "renamed": diff_report["fuzzy_matches"]["count"],
        "format_issues": diff_report["format_inconsistencies"]["count"],
        "compared_against_version": prev_version,
        "compared_against_label": prev_version_entry["label"],
    }

    diff_report["day_summary"] = day_summary
    diff_report["insights"] = insights
    diff_report["schema_mapping"] = manual_mapping
    diff_report_filename = save_series_diff_json(series_id, next_version, diff_report)
    amt_src = request.form.get('amount_source_col') or None
    amt_tgt = request.form.get('amount_target_col') or None
    excel_report_info = store_series_excel_report(
        series_id, series["name"], prev_version_entry["label"], label,
        next_version, diff_report, key_columns, day_summary,
        amount_source_col=amt_src, amount_target_col=amt_tgt,
    )

    version_entry = add_series_version(
        series_id, new_file.filename, df_new, key_columns, diff_summary_meta,
        excel_report_info["report_file"], label=label, data_type=data_type,
    )

    db.upsert_series_metadata(series_id, series["name"], key_columns, user_id=user_id)
    db.upsert_series_version(
        series_id, next_version, label, new_file.filename,
        int(len(df_new)), int(len(df_new.columns)), key_columns, diff_summary_meta,
        excel_report_info["report_file"],
    )
    db.save_row_snapshot(series_id, prev_version, key_columns, df_prev)
    db.save_row_snapshot(series_id, next_version, key_columns, df_new)

    # Append to reconciliation history in datasets table
    from datetime import datetime, timezone
    db.append_reconciliation_history(series_id, {
        "version": next_version,
        "label": label,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "added": diff_summary_meta["added"],
        "deleted": diff_summary_meta["deleted"],
        "updated": diff_summary_meta["updated"],
    })

    return jsonify({
        "series_id": series_id,
        "version": version_entry,
        "compared_against_version": prev_version,
        "key_columns": key_columns,
        "data_type": data_type,
        "report": diff_report,
        "day_summary": day_summary,
        "insights": insights,
        "diff_report_file": diff_report_filename,
        "excel_report_file": excel_report_info["report_file"],
    }), 201


@app.route('/api/series/<series_id>/versions/<int:version>/report', methods=['GET'])
@optional_auth
def series_version_report(series_id, version):
    user_id = getattr(g, 'current_user_id', None)
    series = get_series(series_id)
    if not series:
        return jsonify({"error": "Series not found."}), 404
    if user_id is not None and not getattr(g, 'is_admin', False) and db.is_available():
        owner = db.get_series_owner(series_id)
        if owner is not None and owner != user_id:
            return jsonify({"error": "Access denied."}), 403
    if version == 0:
        return jsonify({"error": "Version 0 is the baseline; nothing to diff against."}), 400
    report = load_series_diff_json(series_id, version)
    if report is None:
        return jsonify({"error": "Report not found for this version."}), 404
    return jsonify({"series_id": series_id, "version": version, "report": report})


# ── Value history (Postgres-backed day-over-day pivot) ────────────────────────

@app.route('/api/series/<series_id>/history', methods=['GET'])
@optional_auth
def series_value_history(series_id):
    if not db.is_available():
        return jsonify({"error": "History requires a connected Postgres database.", "db_connected": False}), 503

    user_id = getattr(g, 'current_user_id', None)
    series = get_series(series_id)
    if not series:
        return jsonify({"error": "Series not found."}), 404
    if user_id is not None and not getattr(g, 'is_admin', False) and db.is_available():
        owner = db.get_series_owner(series_id)
        if owner is not None and owner != user_id:
            return jsonify({"error": "Access denied."}), 403

    only_changed = request.args.get('only_changed', 'true').lower() != 'false'
    history = db.get_value_history(series_id, only_changed=only_changed)
    return jsonify({
        "series_id": series_id,
        "db_connected": True,
        "versions": history["versions"],
        "entries": history["entries"],
    })


# =============================================================================
# AR RECONCILIATION — schema-agnostic, canonical-field-mapped
# =============================================================================
try:
    from ar_column_mapper import _load_synonyms, detect_file_type, load_and_map
    from ar_reconcile import reconcile as ar_reconcile_engine
    from ar_pagination import store_job, get_job, paginate_list, normalize_page_size, DEFAULT_PAGE_SIZE, ALLOWED_PAGE_SIZES

    @app.route('/api/ar/detect-type', methods=['POST'])
    @optional_auth
    def ar_detect_type():
        """Detect whether an uploaded file is Transactional or Master Data."""
        if 'file' not in request.files:
            return jsonify({"error": "Please upload a file as 'file'."}), 400
        uploaded = request.files['file']
        if not allowed_file(uploaded.filename):
            return jsonify({"error": "Allowed file types: csv, xls, xlsx."}), 400
        try:
            df = normalize_dataframe(read_dataframe(uploaded))
        except Exception as exc:
            return jsonify({"error": f"Could not read file: {exc}"}), 400
        synonyms_all = _load_synonyms()
        result = detect_file_type(df, synonyms_all)
        return jsonify(result)

    @app.route('/api/ar/reconcile', methods=['POST'])
    @optional_auth
    def ar_reconcile():
        """AR reconciliation: map columns canonically, then reconcile."""
        if 'source_file' not in request.files or 'target_file' not in request.files:
            return jsonify({"error": "Please upload both source_file and target_file."}), 400
        source_file = request.files['source_file']
        target_file = request.files['target_file']
        if not allowed_file(source_file.filename) or not allowed_file(target_file.filename):
            return jsonify({"error": "Allowed file types: csv, xls, xlsx."}), 400

        def _sanitize(obj):
            """Recursively convert pandas/numpy non-JSON-safe types to Python natives."""
            import pandas as _pd
            import numpy as _np
            if isinstance(obj, dict):
                return {k: _sanitize(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [_sanitize(v) for v in obj]
            if isinstance(obj, _pd.Timestamp):
                return None if _pd.isna(obj) else obj.isoformat()
            if isinstance(obj, float) and _pd.isna(obj):
                return None
            if obj is _pd.NaT:
                return None
            if isinstance(obj, _np.integer):
                return int(obj)
            if isinstance(obj, _np.floating):
                return None if _np.isnan(obj) else float(obj)
            if isinstance(obj, _np.bool_):
                return bool(obj)
            return obj

        try:
            import pandas as _pd
            from io import BytesIO as _BytesIO
            from ar_chunked_read import read_raw_chunked as _read_raw_chunked
            src_content = source_file.read()
            tgt_content = target_file.read()
            # Chunk-based ingestion for large files: CSVs are streamed and
            # concatenated in bounded-size chunks (bounded peak memory)
            # instead of one single parse call; Excel keeps the existing
            # pd.read_excel path (openpyxl already streams the underlying
            # zip/XML, and we must not change how .xlsx files with merged
            # headers etc. were previously read).
            src_raw = _read_raw_chunked(src_content, source_file.filename)
            tgt_raw = _read_raw_chunked(tgt_content, target_file.filename)
        except Exception as exc:
            return jsonify({"error": f"Could not read files: {exc}"}), 400

        synonyms_all = _load_synonyms()
        txn_fields = synonyms_all.get("transactional", {})

        src_overrides = {}
        tgt_overrides = {}
        overrides_raw = request.form.get('overrides')
        if overrides_raw:
            try:
                parsed = json.loads(overrides_raw)
                src_overrides = parsed.get('source', {})
                tgt_overrides = parsed.get('target', {})
            except Exception:
                pass

        try:
            # fuzzy_cutoff is accepted for backward compatibility with any
            # existing frontend still sending it, but is no longer used --
            # the AR column mapper is now fully deterministic (no fuzzy /
            # AI matching of any kind).
            request.form.get('fuzzy_cutoff', None)
            tolerance = float(request.form.get('tolerance', 0.01))
        except ValueError:
            return jsonify({"error": "tolerance must be a number."}), 400

        is_async = request.form.get('async', 'false').lower() in ('true', '1')
        user_id = getattr(g, 'current_user_id', None)

        if is_async:
            job_id = f"job_{os.urandom(8).hex()}"
            src_meta = store_file(source_file.filename, src_raw, "source_upload", user_id=user_id)
            tgt_meta = store_file(target_file.filename, tgt_raw, "target_upload", user_id=user_id)

            db.create_recon_job(
                job_id=job_id,
                user_id=user_id,
                job_type="AR_RECONCILE",
                status="QUEUED_KAFKA",
                source_filename=source_file.filename,
                target_filename=target_file.filename,
                payload_params={"tolerance": tolerance, "overrides": {"source": src_overrides, "target": tgt_overrides}},
            )

            job_payload = {
                "job_id": job_id,
                "user_id": user_id,
                "source_file_id": src_meta["file_id"],
                "target_file_id": tgt_meta["file_id"],
                "overrides": {"source": src_overrides, "target": tgt_overrides},
                "tolerance": tolerance,
                "source_filename": source_file.filename,
                "target_filename": target_file.filename,
            }

            kafka_sent = publish_recon_job(job_id, job_payload)
            publish_audit_event("AR_RECON_QUEUED", {"job_id": job_id, "kafka_sent": kafka_sent}, user_id=user_id)

            return jsonify({
                "job_id": job_id,
                "status": "QUEUED_KAFKA" if kafka_sent else "QUEUED_FALLBACK",
                "message": "Reconciliation job submitted to Kafka workers.",
                "async": True,
            }), 202

        # Source and Target are mapped INDEPENDENTLY onto the same
        # canonical schema -- two separate load_and_map() calls, each
        # running the full Manual Override -> Exact Synonym -> Exact
        # Normalized Name -> Datatype/Sample-Value Validation pipeline.
        src_std, src_mapping, src_report = load_and_map(src_raw, txn_fields, src_overrides)
        tgt_std, tgt_mapping, tgt_report = load_and_map(tgt_raw, txn_fields, tgt_overrides)

        # Required fields must be cleanly MAPPED -- a REVIEW_REQUIRED
        # (ambiguous) or NOT_FOUND field is never silently accepted as
        # mapped, so both statuses block reconciliation the same way.
        def _unresolved_required(report):
            return [r['field'] for r in report
                    if r['status'] != 'MAPPED' and r['field'] in ('TxnNumber', 'Amount')]

        not_found_src = _unresolved_required(src_report)
        not_found_tgt = _unresolved_required(tgt_report)
        if not_found_src or not_found_tgt:
            return jsonify(_sanitize({
                "error": "Required fields not mapped.",
                "not_found_source": not_found_src,
                "not_found_target": not_found_tgt,
                "source_mapping": src_report,
                "target_mapping": tgt_report,
            })), 422

        results = ar_reconcile_engine(src_std, tgt_std, tolerance=tolerance)

        # Result row buckets -- these are what get paginated/lazy-loaded.
        buckets = {
            "matched": results["matched_rows"],
            "disputed": results["mismatch_rows"],
            "unmatched_source": results["only_source_rows"],
            "unmatched_target": results["only_target_rows"],
            "tier2_rows": results["tier2_rows"],
            "duplicate_source_rows": results["duplicate_source_rows"],
            "duplicate_target_rows": results["duplicate_target_rows"],
            "source_exceptions": results["source_exceptions"],
            "target_exceptions": results["target_exceptions"],
        }
        sanitized_buckets = {k: _sanitize(v) for k, v in buckets.items()}
        job_id = store_job(sanitized_buckets, meta={"summary": results["summary"]})

        publish_audit_event("AR_RECON_COMPLETED_SYNC", {"job_id": job_id, "summary": results["summary"]}, user_id=user_id)

        try:
            page_size = normalize_page_size(request.form.get('page_size', DEFAULT_PAGE_SIZE))
        except ValueError:
            page_size = DEFAULT_PAGE_SIZE

        # Server-side pagination: only the first page of each bucket (plus
        # its total count) is returned here -- never the full record set.
        # Additional pages are fetched lazily via GET
        # /api/ar/reconcile/results/<job_id>?bucket=<name>&page=<n>.
        first_pages = {
            name: paginate_list(rows, page=1, page_size=page_size)
            for name, rows in sanitized_buckets.items()
        }

        response = {
            "job_id": job_id,
            "source_mapping": src_report,
            "target_mapping": tgt_report,
            "summary": results["summary"],
            "page_size": page_size,
            "allowed_page_sizes": list(ALLOWED_PAGE_SIZES),
            "results": first_pages,
            "src_report": src_report,
            "tgt_report": tgt_report,
        }
        return jsonify(_sanitize(response))

    @app.route('/api/ar/reconcile/results/<job_id>', methods=['GET'])
    @optional_auth
    def ar_reconcile_results(job_id):
        """Lazy-load a page of one AR reconcile result bucket.
        Query params: bucket (required), page (default 1),
        page_size (default 50; snapped to the nearest of 25/50/100/250),
        sort_by, sort_dir (asc|desc), q (substring filter). Filtering and
        sorting are always applied before the page slice, and only that
        page is ever returned -- the full bucket is never sent to the
        client, however large the reconciliation is."""
        job = get_job(job_id)
        if not job:
            return jsonify({"error": "Unknown or expired job_id. Re-run reconciliation."}), 404

        bucket = request.args.get('bucket')
        if not bucket or bucket not in job["buckets"]:
            return jsonify({"error": f"bucket must be one of {list(job['buckets'].keys())}"}), 400

        try:
            page = int(request.args.get('page', 1))
            page_size = normalize_page_size(request.args.get('page_size', DEFAULT_PAGE_SIZE))
        except ValueError:
            return jsonify({"error": "page and page_size must be integers."}), 400

        result = paginate_list(
            job["buckets"][bucket],
            page=page, page_size=page_size,
            sort_by=request.args.get('sort_by'),
            sort_dir=request.args.get('sort_dir', 'asc'),
            q=request.args.get('q'),
        )
        result["bucket"] = bucket
        result["allowed_page_sizes"] = list(ALLOWED_PAGE_SIZES)
        return jsonify(result)

    # ── Kafka Job Status Endpoints ──────────────────────────────────────────

    @app.route('/api/jobs/<job_id>/status', methods=['GET'])
    @optional_auth
    def job_status(job_id):
        """Returns the real-time execution status of an async Kafka reconciliation job."""
        job_data = db.get_recon_job(job_id)
        if not job_data:
            # Check in-memory pagination store if already finished
            mem_job = get_job(job_id)
            if mem_job:
                return jsonify({
                    "job_id": job_id,
                    "status": "COMPLETED",
                    "progress_pct": 100.0,
                    "result_summary": {"summary": mem_job.get("meta", {}).get("summary")},
                })
            return jsonify({"error": f"Job '{job_id}' not found."}), 404
        return jsonify(job_data)

    @app.route('/api/jobs', methods=['GET'])
    @optional_auth
    def list_jobs():
        """Returns recent reconciliation jobs."""
        user_id = getattr(g, 'current_user_id', None)
        jobs = db.list_recon_jobs_for_user(user_id=user_id, limit=30)
        return jsonify({"jobs": jobs})

except ImportError as _ar_import_err:
    print(f"[ar_reconcile] Skipped: {_ar_import_err}")
# =============================================================================
# END AR RECONCILIATION
# =============================================================================

# =============================================================================
# DUMMY SERVER INTEGRATION — ADDITIVE ONLY
# =============================================================================
# Everything above this block is the ORIGINAL, unmodified reconciliation
# application (auth, uploads, comparison, reports, series/history, chat,
# etc.) — none of it was changed to add this feature.
#
# This registers ONE new, self-contained Blueprint (see
# backend/dummy_integration/routes.py) that exposes:
#
#     POST /api/dummy-integration/source-upload
#     POST /api/dummy-integration/auto-reconcile   <-- powers the
#          "Fetch Target automatically from Dummy Server" checkbox on the
#          Reconcile Over Time screen: the user uploads only a Source file,
#          and this endpoint detects its business key, calls the
#          independent Dummy Server (backend/dummy_server/app.py, run
#          separately on port 9000) for Target data, then runs it through
#          the EXISTING, UNCHANGED difference_summary()/extract_day_summary()
#          comparison engine (fuzzy matching and Format Only detection
#          included) and stores it as a normal Series — so it shows up in
#          "Reconcile Over Time", "Stored Files", and "Reports" exactly
#          like a manual two-file comparison would.
#     GET  /api/dummy-integration/scheduler/status
#     POST /api/dummy-integration/scheduler/trigger
#     GET  /api/dummy-integration/scheduler/last-result
try:
    from dummy_integration.routes import dummy_integration_bp
    app.register_blueprint(dummy_integration_bp)
except Exception as _dummy_integration_import_error:  # pragma: no cover
    # Defensive: if this optional module or its dependencies (SQLAlchemy,
    # requests, python-dotenv) aren't installed, the EXISTING app must still
    # start and work exactly as before — this feature is additive, not required.
    print(f"[dummy_integration] Skipped (not available): {_dummy_integration_import_error}")

# Background scheduler — auto-fetches Target data + reconciles on an
# interval (default every 10 minutes). Guarded the same way as the
# blueprint registration above: never prevents the rest of the app from
# starting if APScheduler isn't installed or the scheduler fails to start.
import os as _os
_is_reloader_process = _os.environ.get("WERKZEUG_RUN_MAIN") == "true"
_debug_mode = _os.environ.get("FLASK_DEBUG", "0") in ("1", "true", "True")

if not _debug_mode or _is_reloader_process:
    try:
        from dummy_integration.scheduler import start_scheduler, stop_scheduler
        _scheduler_instance = start_scheduler()

        import atexit as _atexit
        _atexit.register(stop_scheduler)
    except Exception as _sched_err:
        print(f"[scheduler] Could not start — scheduled reconciliation disabled: {_sched_err}")
# =============================================================================
# END DUMMY SERVER INTEGRATION
# =============================================================================

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
